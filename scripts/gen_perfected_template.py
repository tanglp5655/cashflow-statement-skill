# -*- coding: utf-8 -*-
"""生成《编制现金流量表2024_完善版.xlsx》：
- 表外数据录入页升级为 7 列布局（序号|项目|金额|税率|税率说明|数据来源|说明）
  + 5 个公式自动算行（粉红）+ 顶部警告 + 用户已填数据与数据来源备注
- 资产负债表 / 利润及利润分配表 / 现金流量表：从原模板复制数值与结构
- 编制说明：保留空表
不修改原文件《编制现金流量表2024.xls》。
用法：python scripts/gen_perfected_template.py
"""
import os
import xlrd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

SRC = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "编制现金流量表2024.xls")
DST = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "编制现金流量表2024_完善版.xlsx")

# ================= 表外数据录入页：7 列完善版 =================
# kind: DATA=录入项 / FORMULA=公式自动算（粉红）
# 用户已填数据 + 数据来源备注来自原模板（原 D 列备注）
SECTION1 = [
    ("DATA", "应收票据贴现利息支出", "", "", "选填", None, None, "票据贴现产生的利息（如无贴现可留空）"),
    ("DATA", "★ 支付给职工的工资", 155172.54, "来自于管理费用和销售费用中的工资", "必填", None, None,
     "全年实发工资总额（含奖金津贴）。系统不自动推算，必须手填。"),
    ("DATA", "支付给职工的四金", 41275.90, "来自于管理费用中和销售费用中的社保、医保费", "选填", None, None,
     "社保+公积金等。留空时系统按工资×26.6%自动推算。"),
    ("DATA", "支付给职工的其他福利费", 1644.83, "来自于管理费用和销售费用中的福利费", "选填", None, None,
     "其他职工福利。留空时系统按工资×1.06%自动推算。"),
    ("FORMULA", "⚡ 合计（工资+四金+福利费）", None, "", "自动算", None, None, "=C{r1}+C{r2}+C{r3}", "(公式：其他福利费+四金+工资)"),
    ("DATA", "销项税额", 835728.00, "", "选填", 0.06, "须根据实际税率修改",
     "全年销项税额。留空时系统按收入×6%推算（税率默认6%，须按实际改）。"),
    ("DATA", "进项税额", 0.00, "", "选填", 0.00, "须根据实际税率修改",
     "全年进项税额。留空时系统按成本×6%推算。"),
    ("FORMULA", "⚡ 应交增值税", None, "", "自动算", None, None, "=C{r2}-C{r1}", "(公式：销项税额-进项税额)"),
    ("DATA", "其他各项税", 0.00, "", "选填", None, None,
     "除增值税外各税种（城建税/教育附加等）。留空时系统取利润表营业税金及附加。"),
    ("DATA", "所得税", 1411718.54, "", "选填", 0.25, "须根据实际税率修改",
     "全年所得税。留空时系统取利润表所得税费用。"),
    ("DATA", "管理费用中列支的税金", "", "", "选填", None, None,
     "计入管理费用的税金（房产税、印花税、车船税等）。"),
    ("DATA", "在其他业务支出中列支的税金", "", "", "选填", None, None, "其他业务负担的税金"),
    ("FORMULA", "⚡ 实际应缴纳各项税金合计", None, "", "自动算", None, None,
     "=C{r1}+C{r2}+C{r3}+C{r4}+C{r5}", "(公式：其他业务税金+管理税金+所得税+其他各项税+应交增值税)"),
    ("DATA", "分配股利所支付的现金", "", "", "选填", None, None, "全年向股东分配股利支付的现金"),
    ("DATA", "分配利润所支付的现金", "", "", "选填", None, None, "全年分配利润支付的现金"),
    ("DATA", "利息支出", "", "不抵减利息收入", "选填", None, None,
     "利息支出（不抵减利息收入；利润表财务费用中的利息部分）"),
    ("FORMULA", "⚡ 分配股利、利润或偿付利息所支付的现金", None, "", "自动算", None, None,
     "=C{r1}+C{r2}+C{r3}", "(公式：利息支出+分配利润+分配股利)"),
    ("DATA", "计提坏账准备的比率", 0.005, "可根据实际计提比率修改", "选填", 0.005, "可根据实际计提比率修改",
     "坏账计提比例（默认0.5%，按公司实际改）"),
    ("DATA", "计提的坏账准备", "", "", "选填", None, None, "当年计提坏账（不填为0）"),
    ("DATA", "无形资产摊销", "", "", "选填", None, None, "当年无形资产摊销额"),
    ("DATA", "长期待摊费用摊销", "", "", "选填", None, None, "当年长期待摊费用摊销额"),
    ("DATA", "固定资产报废损失", "", "", "选填", None, None, "当年报废损失（不计入处置现金净额）"),
    ("DATA", "收到投资分红或利润", "", "", "选填", None, None, "当年收到的对外投资分红/利润"),
    ("DATA", "处置固定资产收回的现金净额", "", "", "选填", None, None, "处置固定资产收到的现金净额"),
    ("DATA", "处置无形资产收回的现金净额", "", "", "选填", None, None, "处置无形资产收到的现金净额"),
    ("DATA", "处置其他长期资产收回的现金净额", "", "", "选填", None, None, "处置其他长期资产收到的现金净额"),
    ("FORMULA", "⚡ 处置长期资产现金净额合计", None, "", "自动算", None, None,
     "=C{r1}+C{r2}+C{r3}", "(公式：处置固定资产+处置无形资产+处置其他长期资产)"),
    ("DATA", "收到的其他与投资活动有关的现金", "", "", "选填", None, None, "投资活动其他现金流入"),
    ("DATA", "支付的其他与投资活动有关的现金", "", "", "选填", None, None, "投资活动其他现金流出"),
]

NOTES = [
    "★ = 必填项（系统不会自动推算，必须操作员手填）。⚡ = 自动算公式（粉红阴影，不要填）。其他 = 选填（留空时系统按说明自动推算）。",
    "★ 全年实发工资是唯一必填项；其余留空时系统自动按默认比例推算（详见每个项目的\"说明\"列）。",
    "★ 税率列默认值：销项 6% / 进项 0% / 所得税 25% / 坏账 0.5%，请按公司实际修改。",
    "★ 数据来源列（浅蓝）：建议每个数字注明出处（如\"管理费用明细-工资\"），便于审计追溯。",
    "★ 用引擎跑：python cash_flow_generator.py --bs 资产负债表 --pl 利润表 --extra 本表",
    "★ 本表为完善版（7列布局），原《编制现金流量表2024.xls》保留作公式口径参考。",
    "★ 如果不想手填，可直接用 Python 引擎自动生成（默认模式），再补全勾稽差异较大的项目。",
]


def _resolve_formulas(rows):
    resolved = []
    for i, row in enumerate(rows):
        if row[0] != "FORMULA":
            resolved.append(row)
            continue
        _, name, val, src, req, rate, rate_note, formula_tpl, note = row
        n_above = i
        formula = formula_tpl
        for k in range(1, n_above + 1):
            formula = formula.replace("{r" + str(k) + "}", f"__ROW_{k}__")
        resolved.append((row[0], name, val, src, req, rate, rate_note, formula, note, n_above))
    return resolved


def _copy_sheet_values(wb_src, sh_src, ws_dst):
    """把原 sheet 的数值+结构复制到目标 sheet（不保留公式，取缓存值）。"""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(sh_src.nrows):
        for c in range(sh_src.ncols):
            v = sh_src.cell_value(r, c)
            if v is None or v == "":
                continue
            cell = ws_dst.cell(row=r + 1, column=c + 1, value=v)
            cell.border = border
            if isinstance(v, (int, float)):
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")
    return ws_dst


def build():
    src_wb = xlrd.open_workbook(SRC)
    wb = Workbook()

    # ---- 编制说明（空）----
    ws = wb.active
    ws.title = "编制说明！"

    # ---- 资产负债表 ----
    ws2 = wb.create_sheet("资产负债表")
    _copy_sheet_values(src_wb, src_wb.sheet_by_name("资产负债表"), ws2)

    # ---- 利润及利润分配表 ----
    ws3 = wb.create_sheet("利润及利润分配表")
    _copy_sheet_values(src_wb, src_wb.sheet_by_name("利润及利润分配表"), ws3)

    # ---- 现金流量表 ----
    ws5 = wb.create_sheet("现金流量表")
    _copy_sheet_values(src_wb, src_wb.sheet_by_name("现金流量表"), ws5)

    # ---- 表外数据录入（7列完善版）----
    ws4 = wb.create_sheet("表外数据录入")

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_font = Font(name="宋体", size=14, bold=True)
    head_font = Font(name="宋体", size=11, bold=True)
    normal_font = Font(name="宋体", size=11)
    must_fill = PatternFill("solid", fgColor="FFF2CC")
    head_fill = PatternFill("solid", fgColor="BDD7EE")
    note_fill = PatternFill("solid", fgColor="F2F2F2")
    warn_fill = PatternFill("solid", fgColor="FFE699")
    warn_font = Font(name="宋体", size=11, bold=True, color="C00000")
    formula_fill = PatternFill("solid", fgColor="FFC7CE")
    formula_font = Font(name="宋体", size=11, bold=True, color="9C0006")
    rate_fill = PatternFill("solid", fgColor="E2EFDA")
    source_fill = PatternFill("solid", fgColor="DEEBF7")
    section_fill = PatternFill("solid", fgColor="305496")
    section_font = Font(name="宋体", size=12, bold=True, color="FFFFFF")

    ws4.merge_cells("A1:G1")
    ws4["A1"] = "表 外 数 据 录 入（完善版）"
    ws4["A1"].font = title_font
    ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 28

    ws4.merge_cells("A2:G2")
    ws4["A2"] = ("⚠ 重要提醒：粉红阴影部分已设置为公式，数据自动计算产生，务请不要填入任何数据以免出错！"
                 "只需在「金额」栏白色空格内录入数据。建议在「数据来源」栏填每个数字的出处，便于审计追溯。")
    ws4["A2"].font = warn_font
    ws4["A2"].fill = warn_fill
    ws4["A2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws4.row_dimensions[2].height = 48

    ws4.merge_cells("A3:G3")
    ws4["A3"] = "一、编制现金流量表所需数据录入"
    ws4["A3"].font = section_font
    ws4["A3"].fill = section_fill
    ws4["A3"].alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[3].height = 22

    hdr = ["序号", "项  目", "金  额", "税  率", "税率说明", "数据来源", "说  明"]
    for ci, h in enumerate(hdr, 1):
        c = ws4.cell(row=4, column=ci, value=h)
        c.font = head_font
        c.fill = head_fill
        c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center")

    resolved = _resolve_formulas(SECTION1)
    r = 5
    seq = 0
    for row_def in resolved:
        kind = row_def[0]
        if kind == "DATA":
            _, name, val, src, req, rate, rate_note, note = row_def
            seq += 1
            ws4.cell(row=r, column=1, value=seq).border = border
            ws4.cell(row=r, column=1).alignment = Alignment(horizontal="center")
            cname = ws4.cell(row=r, column=2, value=name)
            cname.border = border
            cname.font = normal_font
            cval = ws4.cell(row=r, column=3)
            if val is not None and val != "":
                cval.value = val
            cval.border = border
            cval.alignment = Alignment(horizontal="right")
            cval.font = normal_font
            crate = ws4.cell(row=r, column=4)
            if rate is not None:
                crate.value = rate
                crate.number_format = "0.00%"
            crate.border = border
            crate.alignment = Alignment(horizontal="center")
            crate.fill = rate_fill
            crate.font = normal_font
            crate_note = ws4.cell(row=r, column=5, value=rate_note or "")
            crate_note.border = border
            crate_note.font = Font(name="宋体", size=9, color="808080")
            crate_note.fill = note_fill
            csrc = ws4.cell(row=r, column=6, value=src or "")
            csrc.border = border
            csrc.fill = source_fill
            csrc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            csrc.font = Font(name="宋体", size=9, italic=True, color="666666")
            cnote = ws4.cell(row=r, column=7, value=note)
            cnote.border = border
            cnote.font = Font(name="宋体", size=9, color="808080")
            cnote.fill = note_fill
            cnote.alignment = Alignment(wrap_text=True, vertical="center")
            if req == "必填":
                cname.fill = must_fill
                cval.fill = must_fill
            r += 1
        else:
            _, name, val, src, req, rate, rate_note, formula_tpl, note, n_above = row_def
            formula = formula_tpl
            for k in range(1, n_above + 1):
                formula = formula.replace(f"__ROW_{k}__", str(r - k))
            ws4.cell(row=r, column=1).border = border
            ws4.cell(row=r, column=1).fill = formula_fill
            cname = ws4.cell(row=r, column=2, value=name)
            cname.border = border
            cname.font = formula_font
            cname.fill = formula_fill
            cval = ws4.cell(row=r, column=3, value=formula)
            cval.border = border
            cval.alignment = Alignment(horizontal="right")
            cval.font = formula_font
            cval.fill = formula_fill
            cval.number_format = "#,##0.00;-#,##0.00;-"
            crate = ws4.cell(row=r, column=4)
            crate.border = border
            crate.fill = formula_fill
            crate_note = ws4.cell(row=r, column=5)
            crate_note.border = border
            crate_note.fill = formula_fill
            csrc = ws4.cell(row=r, column=6)
            csrc.border = border
            csrc.fill = formula_fill
            cnote = ws4.cell(row=r, column=7, value=note)
            cnote.border = border
            cnote.font = Font(name="宋体", size=9, color="9C0006")
            cnote.fill = formula_fill
            cnote.alignment = Alignment(wrap_text=True, vertical="center")
            r += 1

    r += 1
    ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    ws4.cell(row=r, column=1, value="二、编制财务分析表所需数据录入").font = section_font
    ws4.cell(row=r, column=1).fill = section_fill
    ws4.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[r].height = 22
    r += 1
    for ci, h in enumerate(["序号", "项  目", "金  额", "", "", "数据来源", "说  明"], 1):
        c = ws4.cell(row=r, column=ci, value=h)
        c.font = head_font
        c.fill = head_fill
        c.border = border
        c.alignment = Alignment(horizontal="center")
    ws4.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
    r += 1
    ws4.cell(row=r, column=1, value=1).border = border
    ws4.cell(row=r, column=1).alignment = Alignment(horizontal="center")
    cname = ws4.cell(row=r, column=2, value="待摊费用摊销")
    cname.border = border
    cname.font = normal_font
    cval = ws4.cell(row=r, column=3)
    cval.border = border
    cval.alignment = Alignment(horizontal="right")
    ws4.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
    ws4.cell(row=r, column=4).border = border
    ws4.cell(row=r, column=4).fill = note_fill
    csrc = ws4.cell(row=r, column=6)
    csrc.border = border
    csrc.fill = source_fill
    cnote = ws4.cell(row=r, column=7, value="当年待摊费用摊销额（财务分析用）")
    cnote.border = border
    cnote.font = Font(name="宋体", size=9, color="808080")
    cnote.fill = note_fill
    r += 1

    r += 1
    for note in NOTES:
        ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        cell = ws4.cell(row=r, column=1, value="※ " + note)
        cell.font = Font(name="宋体", size=9, color="C00000")
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws4.row_dimensions[r].height = 18
        r += 1

    ws4.column_dimensions["A"].width = 6
    ws4.column_dimensions["B"].width = 34
    ws4.column_dimensions["C"].width = 18
    ws4.column_dimensions["D"].width = 10
    ws4.column_dimensions["E"].width = 18
    ws4.column_dimensions["F"].width = 32
    ws4.column_dimensions["G"].width = 40

    wb.save(DST)
    print("已生成:", DST)


if __name__ == "__main__":
    build()
