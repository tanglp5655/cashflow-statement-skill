# -*- coding: utf-8 -*-
"""
现金流量表自动编制引擎
======================
依据《现金流量表编制说明》与模板《编制现金流量表2024.xls》的公式口径，
根据资产负债表、利润及利润分配表（损益表）及表外数据，自动编制现金流量表。

编制方法：直接法编制主表 + 间接法编制补充资料（净利润调节为经营活动现金流量）。
平衡机制：主表"现金及现金等价物净增加额"与补充资料一致；"收到的其他与经营活动
有关的现金"作为平衡项自动倒挤（与模板公式设置一致）。

输入支持三种格式：
  1) Excel 文件（.xlsx/.xls，兼容模板格式）
  2) CSV 文件（UTF-8 或 GBK 编码）
  3) 粘贴的 Markdown 表格文本（Claude 调用时直接传入）

输出：
  - 现金流量表主表（项目、行次、本期金额、上期金额）
  - 补充资料（间接法调节）
  - 编制附注（各关键项目计算公式与说明）
  - 勾稽关系校验报告

用法（命令行）：
  python cash_flow_generator.py --bs 资产负债表.xlsx --pl 利润表.xlsx [--extra 表外数据.xlsx]
                                [--bs_prior 上年资产负债表.xlsx] [--pl_prior 上年利润表.xlsx]
                                [--out 输出目录] [--fmt excel|markdown|json]
"""
import json
import os
import re
import sys
import csv
import argparse
from copy import deepcopy
from collections import OrderedDict

# ---------------------------------------------------------------------------
# 常量与科目映射（严格按模板《编制现金流量表2024.xls》行次与公式口径）
# ---------------------------------------------------------------------------

# 主表项目行次（会企03表，按模板）
MAIN_ITEMS = [
    # (行次, 项目名称, 类别)
    (1,  "一、经营活动产生的现金流量：", "header"),
    (1,  "  销售商品、提供劳务收到的现金", "op"),
    (3,  "  收到的税费返还", "op"),
    (8,  "  收到的其他与经营活动有关的现金", "op"),
    (9,  "  现金流入小计", "op"),
    (10, "  购买商品、接受劳务支付的现金", "op"),
    (12, "  支付给职工以及为职工支付的现金", "op"),
    (13, "  支付的各项税费", "op"),
    (18, "  支付的其他与经营活动有关的现金", "op"),
    (20, "  现金流出小计", "op"),
    (21, "  经营活动产生的现金流量净额", "op"),
    (22, "二、投资活动产生的现金流量：", "header"),
    (22, "  收回投资所收到的现金", "inv"),
    (23, "  取得投资收益所收到的现金", "inv"),
    (25, "  处置固定资产、无形资产和其他长期资产而收到的现金净额", "inv"),
    (28, "  收到的其他与投资活动有关的现金", "inv"),
    (29, "  现金流入小计", "inv"),
    (30, "  购建固定资产、无形资产和其他长期资产所支付的现金", "inv"),
    (31, "  投资所支付的现金", "inv"),
    (35, "  支付的其他与投资活动有关的现金", "inv"),
    (36, "  现金流出小计", "inv"),
    (37, "  投资活动产生的现金流量净额", "inv"),
    (38, "三、筹资活动产生的现金流量：", "header"),
    (38, "  吸收投资所收到的现金", "fin"),
    (40, "  借款所收到的现金", "fin"),
    (43, "  收到的其他与筹资活动有关的现金", "fin"),
    (44, "  现金流入小计", "fin"),
    (45, "  偿还债务所支付的现金", "fin"),
    (46, "  分配股利、利润或偿付利息所支付的现金", "fin"),
    (52, "  支付的其他与筹资活动有关的现金", "fin"),
    (53, "  现金流出小计", "fin"),
    (54, "  筹资活动产生的现金流量净额", "fin"),
    (55, "四、汇率变动对现金的影响", "fx"),
    (56, "五、现金及现金等价物净增加额", "total"),
]

# 补充资料项目行次（间接法）
SUPP_ITEMS = [
    (57, "  净利润"),
    (58, "  加：计提的资产减值准备"),
    (59, "  固定资产折旧"),
    (60, "  无形资产摊销"),
    (61, "  长期待摊费用摊销"),
    (64, "  待摊费用减少（减：增加）"),
    (65, "  预提费用增加（减：减少）"),
    (66, "  处置固定资产、无形资产和其他长期资产的损失（减：收益）"),
    (67, "  固定资产报废损失"),
    (68, "  财务费用"),
    (69, "  投资损失（减：收益）"),
    (70, "  递延税款贷项（减：借项）"),
    (71, "  存货的减少（减：增加）"),
    (72, "  经营性应收项目的减少（减：增加）"),
    (73, "  经营性应付项目的增加（减：减少）"),
    (74, "  其他"),
    (75, "  经营活动产生的现金流量净额"),
]

CASH_ITEMS = [
    (79, "现金的期末余额"),
    (80, "减：现金的期初余额"),
    (81, "加：现金等价物的期末余额"),
    (82, "减：现金等价物的期初余额"),
    (83, "现金及现金等价物净增加额"),
]

# 模板固定税率/比例（编制说明：须按公司实际情况修改）
DEFAULT_PARAMS = {
    "sale_vat_rate": 0.13,       # 销售含税换算：营业收入 × (1+13%)
    "purchase_vat_rate": 0.17,   # 购货含税换算：营业成本 × (1+17%)（模板用1.17）
    "extra_sale_tax_rate": 0.06, # 表外：销项税额适用税率（模板默认6%，须修改）
    "extra_pur_tax_rate": 0.06,  # 表外：进项税额适用税率（模板默认6%，须修改）
    "four_gold_rate": 0.266,     # 支付给职工的四金比例（模板26.6%）
    "welfare_rate": 0.0106,      # 支付给职工的其他福利费比例（模板1.06%）
    "bad_debt_rate": 0.005,      # 计提坏账准备比率（模板0.5%）
}

# 资产负债表科目规范名（用于匹配用户输入）
BS_KEYS = [
    "货币资金", "短期投资", "应收票据", "应收股利", "应收利息", "应收账款",
    "其他应收款", "预付款项", "应收补贴款", "存货", "待摊费用",
    "一年内到期的长期债券投资", "其他流动资产", "长期股权投资", "长期债权投资",
    "固定资产", "累计折旧", "固定资产减值准备", "工程物资", "在建工程",
    "固定资产清理", "无形资产", "长期待摊费用", "其他长期资产", "递延税款借项",
    "短期借款", "应付票据", "应付账款", "预收款项", "应付职工薪酬",
    "应付福利费", "应付股利", "应交税费", "其他应交款", "其他应付款",
    "预提费用", "预计负债", "一年内到期的长期负债", "其他流动负债",
    "长期借款", "应付债券", "长期应付款", "专项应付款", "其他长期负债",
    "递延税项贷项", "实收资本", "已归还投资", "资本公积", "盈余公积",
    "法定公益金", "未分配利润",
]

# ===== 科目别名映射表（新准则 → 模板科目 / 行业口径 → 模板科目）=====
# 用于适配不同准则、不同行业的报表科目名，使引擎"适应所有公司"
BS_ALIAS = {
    # —— 新收入准则（2017后）—— 合同负债/合同资产已入聚合组求和
    # —— 新金融工具准则 ——
    "交易性金融资产": "短期投资",
    "债权投资": "长期债权投资",
    "其他债权投资": "长期债权投资",
    "其他权益工具投资": "长期股权投资",
    "衍生金融资产": "短期投资",
    "一年内到期的非流动资产": "一年内到期的长期债券投资",
    # —— 新租赁准则 ——
    "使用权资产": "固定资产",
    "租赁负债": "长期应付款",
    # —— 所得税 ——
    "递延所得税资产": "递延税款借项",
    "递延所得税负债": "递延税项贷项",
    # —— 老准则/行业别名 ——
    "预付账款": "预付款项",
    "应收帐款": "应收账款",
    "应付帐款": "应付账款",
    "应付工资": "应付职工薪酬",
    "应付福利费": "应付福利费",
    "应交税金": "应交税费",
    "其他应交款": "其他应交款",
    "股本": "实收资本",
    "库存股": "已归还投资",
    "存货(净额)": "存货",
    "固定资产原值": "固定资产",
    "固定资产(原值)": "固定资产",
    "固定资产净额": "固定资产",
    "商誉": "其他长期资产",
    "开发支出": "无形资产",
    "长期待摊费用": "长期待摊费用",
    "一年内到期的非流动负债": "一年内到期的长期负债",
    # —— 房地产行业（单科目替代）——
    "合同取得成本": "其他流动资产",
    "投资性房地产": "固定资产",
    # —— 建筑行业（单科目替代）——
    "应收票据及应收账款": "应收账款",   # 合并科目：映射其一，防双算
    "应付账款及应付票据": "应付账款",
    "应付工程款": "应付账款",
    # —— 其他行业/通用 ——
    "专项储备": "盈余公积",
    "代理买卖证券款": "其他应付款",
    "预付设备款": "预付款项",
}

# 多科目聚合组：用户把同一模板科目拆成多个明细科目列示时，取值需求和
# （如房地产的开发成本+开发产品=存货；金融的拆入资金并入短期借款）
BS_AGGREGATE = {
    # 制造业/通用存货拆分（老准则分列或个别行业单列）
    "存货": ["原材料", "库存商品", "在产品", "自制半成品", "发出商品",
            "委托加工物资", "周转材料", "低值易耗品", "包装物", "高价周转件",
            # 房地产
            "开发成本", "开发产品", "在建开发产品", "合同履约成本",
            # 农业
            "消耗性生物资产", "农产品", "幼畜及育肥畜", "农业生产成本",
            # 建筑
            "工程施工"],
    "应收票据": ["应收款项融资"],
    "短期借款": ["拆入资金", "向中央银行借款", "卖出回购金融资产款"],
    "其他应收款": ["拆出资金", "买入返售金融资产", "存放中央银行款项",
                    "存放同业款项", "结算备付金", "应收出口退税"],
    "应付账款": ["同业存放", "同业及其他金融机构存放款项", "吸收存款", "客户存款"],
    # 预收类（新老准则并存或行业别名，如万科"预收账款+合同负债"应求和）
    "预收款项": ["合同负债", "预收账款", "预收房款", "预收购房款", "预收售房款", "预收货款"],
    # 应收类（与应收账款并存时应求和）
    "应收账款": ["合同资产", "应收工程款", "应收电费", "应收运费", "租赁应收款"],
    # 长期经营资产（变动均属购建长期资产的现金流出）
    "固定资产": ["油气资产", "井及相关设施", "递耗资产", "生产性生物资产"],
    "在建工程": ["生物性在建工程", "在建船舶"],
    "无形资产": ["矿区权益", "采矿权", "探矿权", "矿业权", "特许经营权",
                 "影视剧版权", "软件著作权", "土地使用权"],
}

# 利润表科目别名
PL_ALIAS = {
    "营业总收入": "营业收入",
    "营业收入(元)": "营业收入",
    "主营营业收入": "营业收入",
    "营业总成本": "营业成本",
    "主营业务税金及附加": "营业税金及附加",
    "营业税金": "营业税金及附加",
    "销售费用": "销售费用",
    "营业费用": "销售费用",
    "管理费用": "管理费用",
    "研发费用": "管理费用",
    "财务费用": "财务费用",
    "投资收益": "投资收益",
    "其中:对联营企业和合营企业的投资收益": "投资收益",
    "营业外收入": "营业外收入",
    "营业外支出": "营业外支出",
    "利润总额": "利润总额",
    "所得税": "所得税费用",
    "所得税费用": "所得税费用",
    "净利润": "净利润",
    "归属于母公司所有者的净利润": "净利润",
    "归母净利润": "净利润",
    "少数股东损益": "少数股东权益",
    "补贴收入": "补贴收入",
    # —— 新准则利润表科目 ——
    "税金及附加": "营业税金及附加",
    "营业税金及附加": "营业税金及附加",
    "资产减值损失": "营业外支出",     # 非付现费用 → 间接法加回（模板用营业外支出近似处置损失）
    "信用减值损失": "营业外支出",     # 同上（金融企业常见）
    "资产处置收益": "营业外收入",     # 近似归类（处置利得非经营现金）
    "资产处置损失": "营业外支出",     # 处置损失非付现 → 间接法加回
    "公允价值变动收益": "投资收益",   # 未实现收益 → 间接法调减（模板投资损失=-投资收益）
}

# 利润表多科目聚合组（用户分列明细时求和）
PL_AGGREGATE = {
    "净利润": ["持续经营净利润", "终止经营净利润"],
    "营业收入": ["主营业务收入", "其他业务收入"],
    "营业成本": ["主营业务成本", "其他业务成本"],
}

# 表外数据科目别名
EXTRA_ALIAS = {
    "支付给职工以及为职工支付的现金": "支付给职工以及为职工支付的现金",
    "支付的各项税费": "支付的各项税费",
    "销售商品、提供劳务收到的现金": "销售商品提供劳务收到的现金",
    "购买商品、接受劳务支付的现金": "购买商品接受劳务支付的现金",
    "收到的其他与经营活动有关的现金": "收到的其他与经营活动有关的现金",
    "支付的其他与经营活动有关的现金": "支付的其他与经营活动有关的现金",
    "购建固定资产、无形资产和其他长期资产支付的现金": "购建固定资产无形资产和其他长期资产支付的现金",
    "购建固定资产、无形资产和其他长期资产所支付的现金": "购建固定资产无形资产和其他长期资产支付的现金",
    "投资所支付的现金": "投资所支付的现金",
    "取得投资收益收到的现金": "取得投资收益所收到的现金",
    "收回投资收到的现金": "收回投资所收到的现金",
    "处置固定资产、无形资产和其他长期资产收回的现金净额": "处置固定资产无形资产和其他长期资产收到的现金净额",
    "借款收到的现金": "借款所收到的现金",
    "取得借款收到的现金": "借款所收到的现金",
    "吸收投资收到的现金": "吸收投资所收到的现金",
    "偿还债务支付的现金": "偿还债务所支付的现金",
    "偿还借款支付的现金": "偿还债务所支付的现金",
    "分配股利、利润或偿付利息支付的现金": "分配股利利润偿付利息现金",
    "汇率变动对现金及现金等价物的影响": "汇率变动对现金的影响",
    "现金及现金等价物净增加额": "现金及现金等价物净增加额",
}

# 表外数据录入项（模板《表外数据录入》工作表）
EXTRA_KEYS = [
    "应收票据贴现利息支出", "支付给职工的工资", "支付给职工的四金",
    "支付给职工的其他福利费", "销项税额", "进项税额", "应交增值税",
    "其他各项税", "所得税", "管理费用中列支的税金",
    "在其他业务支出中列支的税金", "分配股利所支付的现金",
    "分配利润所支付的现金", "利息支出", "计提坏账准备的比率",
    "计提的坏账准备", "无形资产摊销", "长期待摊费用摊销", "固定资产报废损失",
    "收到投资分红或利润", "处置固定资产收回的现金净额",
    "处置无形资产收回的现金净额", "处置其他长期资产收回的现金净额",
    "收到的其他与投资活动有关的现金", "支付的其他与投资活动有关的现金",
]


def _norm(name):
    """科目名称规范化：去空格、全角转半角、统一括号"""
    if name is None:
        return ""
    s = str(name)
    s = s.replace(" ", "").replace("\u3000", "").replace("\t", "")
    s = s.replace("（", "(").replace("）", ")")
    s = s.replace("：", ":").replace("－", "-").replace("—", "-")
    return s.strip()


def norm_key(name):
    """归一化并提取关键匹配词（去除'减'、'其中'等修饰）"""
    s = _norm(name)
    s = s.replace("减:", "").replace("加:", "").replace("其中:", "")
    s = s.replace("收到", "").replace("支付", "")
    s = s.replace("期末", "").replace("期初", "").replace("年初", "")
    s = s.replace("本年", "").replace("上年", "").replace("金额", "")
    s = s.replace("累计数", "")
    return s


# ---------------------------------------------------------------------------
# 输入解析
# ---------------------------------------------------------------------------

def parse_excel_workbook(path):
    """解析 Excel 工作簿为 {工作表名: 单元格列表}，兼容 .xlsx/.xls。
    返回 dict: sheet_name -> list[list]，行优先。"""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xlsx":
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        result = OrderedDict()
        for ws in wb.worksheets:
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append(list(row))
            result[ws.title] = rows
        return result
    elif ext == ".xls":
        import xlrd
        book = xlrd.open_workbook(path)
        result = OrderedDict()
        for sh in book.sheets():
            rows = []
            for r in range(sh.nrows):
                rows.append([sh.cell_value(r, c) for c in range(sh.ncols)])
            result[sh.name] = rows
        return result
    else:
        raise ValueError(f"不支持的 Excel 格式: {ext}（仅支持 .xlsx / .xls）")


def parse_csv(path):
    """解析 CSV（自动探测 UTF-8 / GBK）为 {sheet: rows}"""
    raw = open(path, "rb").read()
    for enc in ("utf-8-sig", "utf-8", "gbk"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = raw.decode("utf-8", errors="replace")
    reader = csv.reader(text.splitlines())
    return {"__csv__": [row for row in reader]}


def parse_markdown_table(text):
    """解析 Markdown 表格（含表头分隔行）为 {sheet: rows}"""
    lines = [ln for ln in text.splitlines() if ln.strip()]
    tables = []
    cur = []
    for ln in lines:
        ln = ln.strip()
        if ln.startswith("|") and ln.endswith("|"):
            cells = [c.strip() for c in ln.strip("|").split("|")]
            # 跳过分隔行 |---|---|
            if all(re.fullmatch(r":?-{2,}:?", c) for c in cells if c):
                continue
            cur.append(cells)
        else:
            if cur:
                tables.append(cur)
                cur = []
    if cur:
        tables.append(cur)
    result = OrderedDict()
    for i, t in enumerate(tables):
        result[f"表{i+1}"] = t
    return result


def load_table(path_or_text):
    """统一入口：根据输入形态解析为 {sheet: rows}。
    支持：dict（已是 {sheet: rows}）、Excel 文件路径、CSV 路径、Markdown 表格文本、CSV 文本。
    """
    if path_or_text is None:
        return None
    # 已是结构化 dict {sheet: rows}
    if isinstance(path_or_text, dict):
        return path_or_text
    s = str(path_or_text).strip()
    if os.path.isfile(s):
        ext = os.path.splitext(s)[1].lower()
        if ext in (".xlsx", ".xls"):
            return parse_excel_workbook(s)
        elif ext == ".csv":
            return parse_csv(s)
        else:
            raise ValueError(f"不支持的文件类型: {ext}")
    if s.startswith("|") or "\n|" in s:
        return parse_markdown_table(s)
    # 尝试当作 CSV 文本
    return {"__text__": [row for row in csv.reader(s.splitlines())]}


# ---------------------------------------------------------------------------
# 数据提取
# ---------------------------------------------------------------------------

def _cell_num(v):
    """转数字，空/非数字返回 None"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").replace("，", "").strip()
    if s in ("", "-", "--", "None"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def extract_named_amounts(rows, name_col, amount_cols, sheet_hint=""):
    """从行数据中提取 {规范科目名: {列名: 金额}}。
    amount_cols: [(列索引, 列标签)]，如 [(2,"年初数"),(3,"期末数")]
    """
    out = OrderedDict()
    for row in rows:
        if name_col >= len(row):
            continue
        name = row[name_col]
        if name is None or str(name).strip() == "":
            continue
        sname = _norm(name)
        if not re.search(r"[\u4e00-\u9fff]", sname):
            continue
        entry = {}
        for ci, label in amount_cols:
            if ci < len(row):
                v = _cell_num(row[ci])
                if v is not None:
                    entry[label] = v
        if entry:
            out[sname] = entry
    return out


def find_bs(rows):
    """在资产负债表的行数据中定位科目/列。返回 (科目列列表, {列标签:列索引列表})。
    支持模板双栏结构：左侧资产 A=科目 B=行次 C=年初 D=期末；右侧负债 E=科目 F=行次 G=年初 H=期末。
    科目列 = "行次"列左边一列。
    """
    header_idx = None
    for i, row in enumerate(rows[:8]):
        joined = "|".join(str(c) if c is not None else "" for c in row[:9])
        if "年初数" in joined and "期末数" in joined:
            header_idx = i
            break
    if header_idx is None:
        return [0], {}
    hrow = rows[header_idx]
    # 定位"行次"列
    line_cols = []
    for ci in range(len(hrow)):
        if _norm(hrow[ci]) == "行次":
            line_cols.append(ci)
    name_cols = []
    for lc in line_cols:
        # 行次列左边一列为科目列
        for ci in range(lc - 1, -1, -1):
            if ci < len(hrow) and str(hrow[ci]).strip():
                name_cols.append(ci)
                break
    # 金额列：每个"年初数"+"期末数"相邻对
    labels = {"年初数": [], "期末数": []}
    for ci in range(len(hrow)):
        lab = _norm(hrow[ci])
        if lab in ("年初数", "期末数", "年初余额", "期末余额"):
            labels[lab].append(ci)
    # 若金额列有两个并列的"年初数"，分别属于左右两栏，按列序配对
    return list(dict.fromkeys(name_cols)) or [0], labels


def parse_balance_sheet(data):
    """解析资产负债表 → {科目: {"年初数":v, "期末数":v}}（含模板两侧）"""
    bs = OrderedDict()
    if data is None:
        return bs
    for sheet_name, rows in data.items():
        name_cols, labels = find_bs(rows)
        if not labels.get("年初数") or not labels.get("期末数"):
            continue
        # 模板双栏：左侧资产栏 (C=年初,D=期末)，右侧负债栏 (G=年初,H=期末)
        # 依据科目列与金额列的相对位置自动配对
        for name_col in name_cols:
            begin_cols = labels["年初数"]
            end_cols = labels["期末数"]
            # 找该科目列右侧最近的一对 (年初数, 期末数)
            pair_begin = min([c for c in begin_cols if c > name_col], default=None)
            pair_end = min([c for c in end_cols if c > name_col], default=None)
            if pair_begin is None or pair_end is None:
                continue
            amount_cols = [(pair_begin, "年初数"), (pair_end, "期末数")]
            extracted = extract_named_amounts(rows, name_col, amount_cols, sheet_name)
            for k, v in extracted.items():
                bs[k] = v
    return bs


def parse_income_statement(data):
    """解析利润及利润分配表 → {科目: {"本年累计数":v, "上年同期数":v}}"""
    pl = OrderedDict()
    if data is None:
        return pl
    for sheet_name, rows in data.items():
        header_idx = None
        for i, row in enumerate(rows[:8]):
            joined = "|".join(str(c) if c is not None else "" for c in row[:5])
            if "本年累计数" in joined or "本年" in joined:
                header_idx = i
                break
        if header_idx is None:
            continue
        hrow = rows[header_idx]
        labels = {}
        for ci in range(len(hrow)):
            lab = _norm(hrow[ci])
            if lab in ("本年累计数", "上年同期数", "本年", "上年"):
                labels[lab] = ci
        # 科目列 = "行次"列左边一列；若无"行次"，取第一个金额列左边最近文本列
        name_col = None
        for ci in range(len(hrow)):
            if _norm(hrow[ci]) == "行次":
                for cj in range(ci - 1, -1, -1):
                    if cj < len(hrow) and str(hrow[cj]).strip():
                        name_col = cj
                        break
                break
        if name_col is None and labels:
            first_amt = min(labels.values())
            for ci in range(first_amt - 1, -1, -1):
                if ci < len(hrow) and str(hrow[ci]).strip():
                    name_col = ci
                    break
        if name_col is None:
            continue
        amount_cols = [(labels[k], k) for k in ("本年累计数", "上年同期数") if k in labels]
        extracted = extract_named_amounts(rows, name_col, amount_cols, sheet_name)
        for k, v in extracted.items():
            pl[k] = v
    return pl


def parse_extra_data(data, bs=None, pl=None, params=None):
    """解析表外数据 → {科目: 金额}，并补全模板中自动计算项。"""
    extra = {}
    if data is not None:
        for sheet_name, rows in data.items():
            for row in rows:
                if len(row) < 3:
                    continue
                name = row[1] if len(row) > 1 else None
                if name is None or str(name).strip() == "":
                    continue
                sname = _norm(name)
                if not re.search(r"[\u4e00-\u9fff]", sname):
                    continue
                v = _cell_num(row[2]) if len(row) > 2 else None
                if v is not None:
                    extra[sname] = v
    params = params or DEFAULT_PARAMS
    # ---- 模板自动计算项 ----
    # 支付给职工的四金 = 工资 × 26.6%（模板）
    wage = extra.get("支付给职工的工资", 0) or 0
    if "支付给职工的四金" not in extra:
        extra["支付给职工的四金"] = round(wage * params["four_gold_rate"], 2)
    if "支付给职工的其他福利费" not in extra:
        extra["支付给职工的其他福利费"] = round(wage * params["welfare_rate"], 2)
    # 销项税额 = 营业收入 × 销项税率（模板 D9，须按实际修改）
    if bs is not None and pl is not None:
        revenue = pl.get("营业收入", {}).get("本年累计数", 0) or 0
        if "销项税额" not in extra and revenue:
            extra["销项税额"] = round(revenue * params["extra_sale_tax_rate"], 2)
        cost = pl.get("营业成本", {}).get("本年累计数", 0) or 0
        inv_begin = bs.get("存货", {}).get("年初数", 0) or 0
        inv_end = bs.get("存货", {}).get("期末数", 0) or 0
        if "进项税额" not in extra and cost:
            extra["进项税额"] = round((cost + inv_end - inv_begin) * params["extra_pur_tax_rate"], 2)
    if "应交增值税" not in extra:
        extra["应交增值税"] = (extra.get("销项税额", 0) or 0) - (extra.get("进项税额", 0) or 0)
    if "所得税" not in extra and pl is not None:
        extra["所得税"] = pl.get("所得税费用", {}).get("本年累计数", 0) or 0
    if "其他各项税" not in extra and pl is not None:
        extra["其他各项税"] = pl.get("营业税金及附加", {}).get("本年累计数", 0) or 0
    # 计提的坏账准备（模板：表外数据"计提的坏账准备"为手工录入项，模板缓存为0）
    # 仅在用户明确录入时使用；未录入则置0，避免偏离模板口径
    if "计提的坏账准备" not in extra:
        extra["计提的坏账准备"] = 0.0
    # 处置固定资产、无形资产和其他长期资产所收回的现金净额合计
    dispo = (extra.get("处置固定资产收回的现金净额", 0) or 0) + \
            (extra.get("处置无形资产收回的现金净额", 0) or 0) + \
            (extra.get("处置其他长期资产收回的现金净额", 0) or 0)
    extra["处置长期资产现金净额合计"] = dispo
    # 分配股利、利润或偿付利息所支付的现金
    extra["分配股利利润偿付利息现金"] = (extra.get("分配股利所支付的现金", 0) or 0) + \
        (extra.get("分配利润所支付的现金", 0) or 0) + (extra.get("利息支出", 0) or 0)
    return extra


# ---------------------------------------------------------------------------
# 科目取值工具
# ---------------------------------------------------------------------------

class SheetData:
    """封装资产负债表/利润表，提供 get() 便捷方法"""

    def __init__(self, mapping, kind):
        self.mapping = mapping   # {规范名: {col: value}}
        self.kind = kind

    def get(self, name, col):
        """按规范科目名取数；自动兼容别名"""
        if not self.mapping:
            return 0.0
        key = self._find(name)
        if key is None:
            return 0.0
        return self.mapping[key].get(col, 0.0) or 0.0

    def _find(self, name):
        target = norm_key(name)
        if target in self.mapping:
            return target
        # 模糊：包含匹配
        for k in self.mapping:
            if target and (target in k or k in target):
                return k
        return None


def _is_alias_key(k, aliases):
    """判断键 k 是否本身就是某个别名（如"应收票据及应收账款"）。
    模糊匹配时跳过这些键，防止合并科目被多个模板科目同时取到（双算）。"""
    nk = _norm(k)
    return bool(nk) and any(_norm(a) == nk for a in aliases)


def _match_keys(keys, name, aliases, aggregate):
    """按科目名查找所有匹配的用户键（支持聚合求和）。
    策略：①精确 → ②聚合组（多明细科目求和，如开发成本+开发产品=存货）
         → ③别名替代（仅未命中时兜底，如投资性房地产→固定资产）
         → ④模糊包含（跳过别名键，取包含匹配中最短者）。
    返回匹配到的键列表（每个键只会被消费一次）。
    """
    if not keys:
        return []
    keys = list(keys)
    matched = []
    used = set()
    target_norm = _norm(name)
    # ① 精确
    if name in keys:
        matched.append(name)
        used.add(name)
    # ② 聚合组：构成同一模板科目的多个明细键求和（无论精确是否命中）
    for member in aggregate.get(name, []):
        for k in keys:
            if k in used:
                continue
            if _norm(k) == _norm(member):
                matched.append(k)
                used.add(k)
    # ③ 别名替代：仅当 name 尚未被精确/聚合命中时兜底（避免投资性房地产并入固定资产类双算）
    if not matched:
        for alias, canon in aliases.items():
            if canon and _norm(canon) == target_norm:
                for k in keys:
                    if _norm(k) == _norm(alias):
                        matched.append(k)
                        return matched  # 单科目替代，取第一个
    # ④ 模糊包含（兜底，跳过别名键）
    if not matched:
        t = norm_key(name)
        best_key, best_len = None, None
        for k in keys:
            if _is_alias_key(k, aliases):
                continue
            kk = norm_key(k)
            if not kk:
                continue
            if kk == t:
                best_key = k
                break
            if t and (t in kk or kk in t):
                if best_len is None or len(kk) < best_len:
                    best_key, best_len = k, len(kk)
        if best_key is not None:
            matched.append(best_key)
    return matched


def _bs(bs, name, col):
    """按科目名取数：精确 → 聚合组（多明细求和）→ 别名替代 → 模糊兜底"""
    if not bs:
        return 0.0
    total = 0.0
    for k in _match_keys(bs.keys(), name, BS_ALIAS, BS_AGGREGATE):
        total += bs[k].get(col, 0.0) or 0.0
    return total


def _pl(pl, name, col="本年累计数"):
    """利润表取值：精确 → 聚合组 → 别名替代 → 模糊兜底（含"一、""减："前缀）"""
    if not pl:
        return 0.0
    total = 0.0
    for k in _match_keys(pl.keys(), name, PL_ALIAS, PL_AGGREGATE):
        total += pl[k].get(col, 0.0) or 0.0
    return total


# ---------------------------------------------------------------------------
# 核心计算引擎（严格按模板公式）
# ---------------------------------------------------------------------------

def compute_cash_flow(bs, pl, extra, params=None):
    """计算现金流量表。
    bs:   {规范科目: {"年初数":v, "期末数":v}}
    pl:   {规范科目: {"本年累计数":v, "上年同期数":v}}
    extra:{规范科目: 金额}
    返回: dict，含 main/supp/cash/adjust/notes/validations
    """
    params = params or DEFAULT_PARAMS
    notes = []          # 编制附注
    adjustments = []    # 负数调整记录

    def bs_get(name, col):
        return _bs(bs, name, col)

    def pl_get(name, col="本年累计数"):
        return _pl(pl, name, col)

    def extra_get(name):
        if name in extra and extra[name]:
            return float(extra[name])
        # 别名/标点匹配（如"销售商品、提供劳务收到的现金"→"销售商品提供劳务收到的现金"）
        target = _norm(name)
        # 先试去掉中文标点的精确匹配
        t2 = target.replace("、", "").replace("，", "").replace(",", "").replace("。", "")
        for k, v in extra.items():
            kk = _norm(k).replace("、", "").replace("，", "").replace(",", "").replace("。", "")
            if kk == t2 and v:
                return float(v)
        for alias, canon in EXTRA_ALIAS.items():
            # 引擎查 canon（如"借款所收到的现金"），用户在 extra 里填的是别名（如"取得借款收到的现金"）
            if canon and _norm(canon) == target and alias in extra and extra[alias]:
                return float(extra[alias])
        return 0.0

    # ============ 表外派生：实际应缴纳各项税金合计 ============
    # 模板：C16 = (应交增值税+其他各项税+所得税+管理税金+其他业务税金)
    #        - (应交税费期末-年初) - (其他应交款期末-年初)
    vat_pay = extra_get("应交增值税")
    other_tax = extra_get("其他各项税")
    income_tax = extra_get("所得税")
    mgmt_tax = extra_get("管理费用中列支的税金")
    other_biz_tax = extra_get("在其他业务支出中列支的税金")
    tax_payable_inc = (bs_get("应交税费", "期末数") + bs_get("其他应交款", "期末数")) \
                    - (bs_get("应交税费", "年初数") + bs_get("其他应交款", "年初数"))
    actual_tax_paid = vat_pay + other_tax + income_tax + mgmt_tax + other_biz_tax - tax_payable_inc
    extra["实际应缴纳各项税金合计"] = actual_tax_paid
    notes.append("支付的各项税费 = 应交增值税(%s) + 其他各项税(%s) + 所得税(%s) + 管理税金(%s)"
                 " + 其他业务税金(%s) - 应交税费及应交款增加(%s)"
                 % (fmt(vat_pay), fmt(other_tax), fmt(income_tax), fmt(mgmt_tax),
                    fmt(other_biz_tax), fmt(tax_payable_inc)))

    # ============ 主表：经营活动 ============
    # 1. 销售商品、提供劳务收到的现金
    #    模板C6 = 营业收入×1.13 + (应收票据年初-期末) + (应收账款年初-期末)
    #            + (预收款项期末-年初) - 应收票据贴现利息支出
    #    高精度模式：用户提供真实值则直接覆盖（从现金流量表主表抄数，最准）
    revenue = pl_get("营业收入")
    _sale_override = extra_get("销售商品提供劳务收到的现金")
    if _sale_override:
        sale_cash = float(_sale_override)
        notes.append("销售商品、提供劳务收到的现金：使用用户提供的现金流量表主表真实值(%s)"
                     % fmt(sale_cash))
    else:
        sale_cash = revenue * (1 + params["sale_vat_rate"]) \
            + (bs_get("应收票据", "年初数") - bs_get("应收票据", "期末数")) \
            + (bs_get("应收账款", "年初数") - bs_get("应收账款", "期末数")) \
            + (bs_get("预收款项", "期末数") - bs_get("预收款项", "年初数")) \
            - extra_get("应收票据贴现利息支出")
        notes.append("销售商品、提供劳务收到的现金 = 营业收入(%s)×(1+%s%%) + (应收票据年初-%s期末:%s)"
                     " + (应收账款年初-期末:%s) + (预收款项期末-年初:%s) - 贴现利息(%s)"
                     % (fmt(revenue), int(params["sale_vat_rate"] * 100),
                        fmt(bs_get("应收票据", "年初数")), fmt(bs_get("应收票据", "期末数")),
                        fmt(bs_get("应收账款", "年初数") - bs_get("应收账款", "期末数")),
                        fmt(bs_get("预收款项", "期末数") - bs_get("预收款项", "年初数")),
                        fmt(extra_get("应收票据贴现利息支出"))))

    # 2. 收到的税费返还 = (应收补贴款年初-期末) + 补贴收入（可真实值覆盖）
    _taxrefund_override = extra_get("收到的税费返还")
    if _taxrefund_override:
        tax_refund = float(_taxrefund_override)
        notes.append("收到的税费返还：使用用户提供的现金流量表主表真实值(%s)" % fmt(tax_refund))
    else:
        tax_refund = (bs_get("应收补贴款", "年初数") - bs_get("应收补贴款", "期末数")) \
            + pl_get("补贴收入")

    # 3. 购买商品、接受劳务支付的现金
    #    模板C10 = (营业成本 + 存货期末-年初)×1.17 + (应付票据年初-期末)
    #            + (应付账款年初-期末) + (预付款项期末-年初)
    _purchase_override = extra_get("购买商品接受劳务支付的现金")
    cost = pl_get("营业成本")
    inv_change = bs_get("存货", "期末数") - bs_get("存货", "年初数")
    if _purchase_override:
        purchase_cash = float(_purchase_override)
        notes.append("购买商品、接受劳务支付的现金：使用用户提供的现金流量表主表真实值(%s)"
                     % fmt(purchase_cash))
    else:
        purchase_cash = (cost + inv_change) * (1 + params["purchase_vat_rate"]) \
            + (bs_get("应付票据", "年初数") - bs_get("应付票据", "期末数")) \
            + (bs_get("应付账款", "年初数") - bs_get("应付账款", "期末数")) \
            + (bs_get("预付款项", "期末数") - bs_get("预付款项", "年初数"))
        notes.append("购买商品、接受劳务支付的现金 = (营业成本(%s)+存货增加(%s))×(1+%s%%)"
                     " + (应付票据年初-期末:%s) + (应付账款年初-期末:%s) + (预付款项期末-年初:%s)"
                     % (fmt(cost), fmt(inv_change), int(params["purchase_vat_rate"] * 100),
                        fmt(bs_get("应付票据", "年初数") - bs_get("应付票据", "期末数")),
                        fmt(bs_get("应付账款", "年初数") - bs_get("应付账款", "期末数")),
                        fmt(bs_get("预付款项", "期末数") - bs_get("预付款项", "年初数"))))

    # 4. 支付给职工以及为职工支付的现金 = 工资+四金+其他福利费（可真实值覆盖）
    _wage_override = extra_get("支付给职工以及为职工支付的现金")
    wage_paid = extra_get("支付给职工的工资") + extra_get("支付给职工的四金") \
        + extra_get("支付给职工的其他福利费")
    if _wage_override:
        wage_paid = float(_wage_override)
        notes.append("支付给职工以及为职工支付的现金：使用用户提供的现金流量表主表真实值(%s)"
                     % fmt(wage_paid))

    # 5. 支付的各项税费 = 表外"实际应缴纳各项税金合计"（可真实值覆盖）
    _tax_override = extra_get("支付的各项税费")
    tax_paid = actual_tax_paid if not _tax_override else float(_tax_override)
    if _tax_override:
        notes.append("支付的各项税费：使用用户提供的现金流量表主表真实值(%s)" % fmt(tax_paid))

    # 6. 支付的其他与经营活动有关的现金（模板为手工/平衡区，默认0，可覆盖）
    other_op_paid = extra_get("支付的其他与经营活动有关的现金")

    op_inflow_sub = sale_cash + tax_refund
    op_outflow_sub = purchase_cash + wage_paid + tax_paid + other_op_paid

    # 7. 投资活动
    # 收回投资所收到的现金（手工，默认0）
    inv_retrieve = extra_get("收回投资所收到的现金")
    # 取得投资收益所收到的现金 = 收到投资分红或利润
    inv_income = extra_get("收到投资分红或利润")
    # 处置固定资产、无形资产和其他长期资产而收到的现金净额
    inv_dispose = extra_get("处置长期资产现金净额合计")
    # 收到的其他与投资活动有关的现金
    inv_other_in = extra_get("收到的其他与投资活动有关的现金")
    # 购建固定资产、无形资产和其他长期资产所支付的现金
    #    模板C22 = (固定资产期末-年初)+(工程物资期末-年初)+(在建工程期末-年初)
    #            +(无形资产及其他资产合计期末-年初)
    #    高精度模式：用户提供真实值则直接覆盖
    _fa_override = extra_get("购建固定资产无形资产和其他长期资产支付的现金")
    if _fa_override:
        fa_purchase = float(_fa_override)
        notes.append("购建固定资产、无形资产和其他长期资产所支付的现金：使用用户提供真实值(%s)"
                     % fmt(fa_purchase))
    else:
        fa_purchase = (bs_get("固定资产", "期末数") - bs_get("固定资产", "年初数")) \
            + (bs_get("工程物资", "期末数") - bs_get("工程物资", "年初数")) \
            + (bs_get("在建工程", "期末数") - bs_get("在建工程", "年初数")) \
            + (bs_get("无形资产", "期末数") - bs_get("无形资产", "年初数")) \
            + (bs_get("长期待摊费用", "期末数") - bs_get("长期待摊费用", "年初数")) \
            + (bs_get("其他长期资产", "期末数") - bs_get("其他长期资产", "年初数"))
        notes.append("购建固定资产、无形资产和其他长期资产所支付的现金 = 固定资产增加(%s)"
                     " + 工程物资增加(%s) + 在建工程增加(%s) + 无形资产及长期资产增加(%s)"
                     % (fmt(bs_get("固定资产", "期末数") - bs_get("固定资产", "年初数")),
                        fmt(bs_get("工程物资", "期末数") - bs_get("工程物资", "年初数")),
                        fmt(bs_get("在建工程", "期末数") - bs_get("在建工程", "年初数")),
                        fmt((bs_get("无形资产", "期末数") - bs_get("无形资产", "年初数"))
                            + (bs_get("长期待摊费用", "期末数") - bs_get("长期待摊费用", "年初数"))
                            + (bs_get("其他长期资产", "期末数") - bs_get("其他长期资产", "年初数")))))
    # 投资所支付的现金 = 短期投资增加 + 长期投资增加（可真实值覆盖）
    _invpay_override = extra_get("投资所支付的现金")
    if _invpay_override:
        inv_pay = float(_invpay_override)
        notes.append("投资所支付的现金：使用用户提供真实值(%s)" % fmt(inv_pay))
    else:
        inv_pay = (bs_get("短期投资", "期末数") - bs_get("短期投资", "年初数")) \
            + (bs_get("长期股权投资", "期末数") - bs_get("长期股权投资", "年初数")) \
            + (bs_get("长期债权投资", "期末数") - bs_get("长期债权投资", "年初数"))
    # 支付的其他与投资活动有关的现金
    inv_other_pay = extra_get("支付的其他与投资活动有关的现金")

    inv_inflow_sub = inv_retrieve + inv_income + inv_dispose + inv_other_in
    inv_outflow_sub = fa_purchase + inv_pay + inv_other_pay

    # 8. 筹资活动
    # 吸收投资所收到的现金 = 实收资本（股本）变动 + 资本公积变动（模板 C28=H34-G34 仅实收资本，
    # 扩展含资本公积以适配定增/H股上市等溢价发行场景——溢价进资本公积，可真实值覆盖）
    _absorb_override = extra_get("吸收投资所收到的现金")
    if _absorb_override:
        fin_absorb = float(_absorb_override)
        notes.append("吸收投资所收到的现金：使用用户提供真实值(%s)" % fmt(fin_absorb))
    else:
        fin_absorb = (bs_get("实收资本", "期末数") - bs_get("实收资本", "年初数")) \
            + (bs_get("资本公积", "期末数") - bs_get("资本公积", "年初数"))
    # 借款所收到的现金 = 短期借款变动 + 长期负债合计变动（模板 C29=H6+H27-G6-G27，可真实值覆盖）
    _borrow_override = extra_get("借款所收到的现金")
    if _borrow_override:
        fin_borrow = float(_borrow_override)
        notes.append("借款所收到的现金：使用用户提供真实值(%s)" % fmt(fin_borrow))
    else:
        # 长期负债合计 = 长期借款+应付债券+长期应付款+专项应付款+其他长期负债
        fin_borrow = (bs_get("短期借款", "期末数") - bs_get("短期借款", "年初数")) \
            + (bs_get("长期借款", "期末数") - bs_get("长期借款", "年初数")) \
            + (bs_get("应付债券", "期末数") - bs_get("应付债券", "年初数")) \
            + (bs_get("长期应付款", "期末数") - bs_get("长期应付款", "年初数")) \
            + (bs_get("专项应付款", "期末数") - bs_get("专项应付款", "年初数")) \
            + (bs_get("其他长期负债", "期末数") - bs_get("其他长期负债", "年初数"))
    fin_other_in = extra_get("收到的其他与筹资活动有关的现金")
    fin_repay = extra_get("偿还债务所支付的现金")
    fin_dividend = extra_get("分配股利利润偿付利息现金")
    fin_other_pay = extra_get("支付的其他与筹资活动有关的现金")

    fin_inflow_sub = fin_absorb + fin_borrow + fin_other_in
    fin_outflow_sub = fin_repay + fin_dividend + fin_other_pay

    # 9. 汇率变动
    fx_change = extra_get("汇率变动对现金的影响")

    # ============ 负数调整（编制说明第四条） ============
    # 模板逻辑：先确定各项金额，再做负数修正（流入负数移入对应流出项），
    # 最后以"收到的其他与经营活动有关的现金"为平衡项倒挤，保证主表自动平衡。
    def fix_negative():
        """主表现金流入栏不得为负数；若出现则移入对应流出栏（编制说明第四条）"""
        nonlocal sale_cash, tax_refund, inv_retrieve, inv_income, \
            inv_dispose, inv_other_in, fin_absorb, fin_borrow, fin_other_in, \
            purchase_cash, wage_paid, tax_paid, other_op_paid, fa_purchase, \
            inv_pay, inv_other_pay, fin_repay, fin_dividend, fin_other_pay

        # 销售商品收到现金为负 → 移至购买商品支付
        if sale_cash < 0:
            adjustments.append("“销售商品、提供劳务收到的现金”为负(%s)，已按正数(%s)转入"
                               "“购买商品、接受劳务支付的现金”" % (fmt(sale_cash), fmt(-sale_cash)))
            purchase_cash += -sale_cash
            sale_cash = 0.0
        # 投资所支付的现金为负 → 调整至收回投资所收到的现金
        if inv_pay < 0:
            adjustments.append("“投资所支付的现金”为负(%s)，已转入“收回投资所收到的现金”"
                               % fmt(inv_pay))
            inv_retrieve += -inv_pay
            inv_pay = 0.0
        # 购建固定资产支付为负 → 调整至处置固定资产收到的现金净额
        if fa_purchase < 0:
            adjustments.append("“购建固定资产、无形资产和其他长期资产所支付的现金”为负(%s)，"
                               "已转入“处置固定资产、无形资产和其他长期资产收到的现金净额”"
                               % fmt(fa_purchase))
            inv_dispose += -fa_purchase
            fa_purchase = 0.0
        # 借款所收到的现金为负 → 调整至偿还债务所支付的现金
        if fin_borrow < 0:
            adjustments.append("“借款所收到的现金”为负(%s)，已转入“偿还债务所支付的现金”"
                               % fmt(fin_borrow))
            fin_repay += -fin_borrow
            fin_borrow = 0.0
        # 吸收投资收到现金为负 → 调整至支付的其他与筹资活动有关的现金
        if fin_absorb < 0:
            adjustments.append("“吸收投资所收到的现金”为负(%s)，已转入“支付的其他与筹资活动有关的现金”"
                               % fmt(fin_absorb))
            fin_other_pay += -fin_absorb
            fin_absorb = 0.0
        # 分配股利利润为负 → 调整至借款所收到的现金
        if fin_dividend < 0:
            adjustments.append("“分配股利、利润或偿付利息所支付的现金”为负(%s)，已转入“借款所收到的现金”"
                               % fmt(fin_dividend))
            fin_borrow += -fin_dividend
            fin_dividend = 0.0
        # 其他流入项为负 → 对应流出项
        for nm, val, tgt in [
            ("收回投资所收到的现金", inv_retrieve, None),
            ("取得投资收益所收到的现金", inv_income, None),
            ("处置固定资产收到的现金净额", inv_dispose, None),
        ]:
            if val < 0:
                adjustments.append("“%s”为负(%s)，已转正处理" % (nm, fmt(val)))

    fix_negative()

    # ============ 汇总（负数修正后） ============
    op_inflow_det = sale_cash + tax_refund
    op_outflow = purchase_cash + wage_paid + tax_paid + other_op_paid
    inv_inflow = inv_retrieve + inv_income + inv_dispose + inv_other_in
    inv_outflow = fa_purchase + inv_pay + inv_other_pay
    fin_inflow = fin_absorb + fin_borrow + fin_other_in
    fin_outflow = fin_repay + fin_dividend + fin_other_pay
    inv_net = inv_inflow - inv_outflow
    fin_net = fin_inflow - fin_outflow

    # ============ 平衡项：收到的其他与经营活动有关的现金 ============
    # 高精度模式：用户提供真实值则直接采用，不再倒挤（此时经营净额 = 各项真实值之和）
    _other_op_in_override = extra_get("收到的其他与经营活动有关的现金")
    if _other_op_in_override:
        other_op_in = float(_other_op_in_override)
        notes.append("收到的其他与经营活动有关的现金：使用用户提供的现金流量表主表真实值(%s)"
                     "（高精度模式：经营活动净额由各真实项求和得出，不再以货币资金变动倒挤）"
                     % fmt(other_op_in))
        # 高精度模式下现金净增加额仍按货币资金+现金等价物口径计算（仅用于勾稽校验参考）
        cash_end = bs_get("货币资金", "期末数")
        cash_begin = bs_get("货币资金", "年初数")
        eqv_end = extra_get("现金等价物期末余额")
        eqv_begin = extra_get("现金等价物期初余额")
        cash_increase = (cash_end - cash_begin) + (eqv_end - eqv_begin)
    else:
        # 模板C8 = 现金净增加额(F38) + 经营流出小计(C14) - 销售现金(C6) - 税费返还(C7)
        #        - 投资净额(C26) - 筹资净额(C36) - 汇率(C37)
        cash_end = bs_get("货币资金", "期末数")
        cash_begin = bs_get("货币资金", "年初数")
        eqv_end = extra_get("现金等价物期末余额")
        eqv_begin = extra_get("现金等价物期初余额")
        cash_increase = (cash_end - cash_begin) + (eqv_end - eqv_begin)

        other_op_in = cash_increase + op_outflow - sale_cash - tax_refund \
            - inv_net - fin_net - fx_change
        notes.append("收到的其他与经营活动有关的现金（平衡项）= 现金及现金等价物净增加额(%s)"
                     " + 经营流出小计(%s) - 销售商品收到现金(%s) - 收到税费返还(%s)"
                     " - 投资活动净额(%s) - 筹资活动净额(%s) - 汇率影响(%s)"
                     % (fmt(cash_increase), fmt(op_outflow), fmt(sale_cash),
                        fmt(tax_refund), fmt(inv_net), fmt(fin_net), fmt(fx_change)))

    # 平衡项若为负 → 按编制说明转入"支付的其他与经营活动有关的现金"
    # 平衡项若为负 → 按编制说明转入"支付的其他与经营活动有关的现金"
    # （高精度模式/用户已提供真实值时不做该调整，保留真实值）
    if other_op_in < 0 and not _other_op_in_override:
        adjustments.append("“收到的其他与经营活动有关的现金”（平衡项）为负(%s)，"
                           "已按正数(%s)转入“支付的其他与经营活动有关的现金”"
                           % (fmt(other_op_in), fmt(-other_op_in)))
        other_op_paid += -other_op_in
        other_op_in = 0.0

    op_inflow = sale_cash + tax_refund + other_op_in
    op_outflow = purchase_cash + wage_paid + tax_paid + other_op_paid
    op_net = op_inflow - op_outflow
    inv_inflow = inv_retrieve + inv_income + inv_dispose + inv_other_in
    inv_outflow = fa_purchase + inv_pay + inv_other_pay
    inv_net3 = inv_inflow - inv_outflow
    fin_inflow = fin_absorb + fin_borrow + fin_other_in
    fin_outflow = fin_repay + fin_dividend + fin_other_pay
    fin_net3 = fin_inflow - fin_outflow
    total_net = op_net + inv_net3 + fin_net3 + fx_change

    # ============ 补充资料（间接法） ============
    net_profit = pl_get("净利润")
    bad_debt = extra_get("计提的坏账准备")
    dep = bs_get("累计折旧", "期末数") - bs_get("累计折旧", "年初数")
    intang_amort = extra_get("无形资产摊销")
    lta_amort = extra_get("长期待摊费用摊销")
    prepaid_dec = bs_get("待摊费用", "年初数") - bs_get("待摊费用", "期末数")
    accrued_inc = bs_get("预提费用", "期末数") - bs_get("预提费用", "年初数")
    dispose_loss = pl_get("营业外支出")  # 模板用营业外支出近似处置损失
    scrap_loss = extra_get("固定资产报废损失")
    fin_expense = extra_get("利息支出")  # 模板：财务费用=利息支出（表外）
    inv_loss = -pl_get("投资收益")
    deferred_tax = (bs_get("递延税项贷项", "期末数") - bs_get("递延税项贷项", "年初数")) \
        - (bs_get("递延税款借项", "期末数") - bs_get("递延税款借项", "年初数"))
    inv_dec = bs_get("存货", "年初数") - bs_get("存货", "期末数")
    # 经营性应收项目减少
    recv_items = ["应收票据", "应收股利", "应收利息", "应收账款", "其他应收款",
                  "预付款项", "应收补贴款"]
    recv_begin = sum(bs_get(i, "年初数") for i in recv_items)
    recv_end = sum(bs_get(i, "期末数") for i in recv_items)
    recv_dec = recv_begin - recv_end
    # 经营性应付项目增加
    pay_items = ["应付票据", "应付账款", "预收款项", "应付职工薪酬", "应付福利费",
                 "应付股利", "应交税费", "其他应交款", "其他应付款", "预提费用"]
    pay_begin = sum(bs_get(i, "年初数") for i in pay_items)
    pay_end = sum(bs_get(i, "期末数") for i in pay_items)
    pay_inc = pay_end - pay_begin

    supp_parts = {
        "净利润": net_profit,
        "计提的资产减值准备": bad_debt,
        "固定资产折旧": dep,
        "无形资产摊销": intang_amort,
        "长期待摊费用摊销": lta_amort,
        "待摊费用减少": prepaid_dec,
        "预提费用增加": accrued_inc,
        "处置长期资产损失": dispose_loss,
        "固定资产报废损失": scrap_loss,
        "财务费用": fin_expense,
        "投资损失": inv_loss,
        "递延税款贷项": deferred_tax,
        "存货的减少": inv_dec,
        "经营性应收项目的减少": recv_dec,
        "经营性应付项目的增加": pay_inc,
    }
    # 模板："其他"为倒挤项，保证补充资料经营净额 = 主表经营净额
    other_supp = op_net - sum(supp_parts.values())
    supp_parts["其他"] = other_supp
    supp_op_net = sum(supp_parts.values())

    # ============ 勾稽校验 ============
    validations = []
    bs_assets = bs_get("固定资产", "年初数")  # 占位，仅用于构建校验
    # 校验1：资产负债表平衡
    asset_items = ["货币资金", "短期投资", "应收票据", "应收股利", "应收利息", "应收账款",
                   "其他应收款", "预付款项", "应收补贴款", "存货", "待摊费用",
                   "一年内到期的长期债券投资", "其他流动资产", "长期股权投资",
                   "长期债权投资", "固定资产", "工程物资", "在建工程", "固定资产清理",
                   "无形资产", "长期待摊费用", "其他长期资产", "递延税款借项"]
    total_assets_begin = sum(bs_get(i, "年初数") for i in asset_items) \
        - bs_get("累计折旧", "年初数") - bs_get("固定资产减值准备", "年初数")
    total_assets_end = sum(bs_get(i, "期末数") for i in asset_items) \
        - bs_get("累计折旧", "期末数") - bs_get("固定资产减值准备", "期末数")
    liab_eq_items = ["短期借款", "应付票据", "应付账款", "预收款项", "应付职工薪酬",
                     "应付福利费", "应付股利", "应交税费", "其他应交款", "其他应付款",
                     "预提费用", "预计负债", "一年内到期的长期负债", "其他流动负债",
                     "长期借款", "应付债券", "长期应付款", "专项应付款", "其他长期负债",
                     "递延税项贷项", "实收资本", "资本公积", "盈余公积", "未分配利润"]
    total_liab_eq_begin = sum(bs_get(i, "年初数") for i in liab_eq_items) \
        - bs_get("已归还投资", "年初数")
    total_liab_eq_end = sum(bs_get(i, "期末数") for i in liab_eq_items) \
        - bs_get("已归还投资", "期末数")
    bs_ok_begin = abs(total_assets_begin - total_liab_eq_begin) < 0.01
    bs_ok_end = abs(total_assets_end - total_liab_eq_end) < 0.01
    validations.append({
        "name": "资产负债表勾稽（资产=负债+权益）",
        "detail": "年初：资产(%s) vs 负债权益(%s)%s；期末：资产(%s) vs 负债权益(%s)%s"
                  % (fmt(total_assets_begin), fmt(total_liab_eq_begin),
                     " ✓" if bs_ok_begin else " ✗ 差异" + fmt(abs(total_assets_begin - total_liab_eq_begin)),
                     fmt(total_assets_end), fmt(total_liab_eq_end),
                     " ✓" if bs_ok_end else " ✗ 差异" + fmt(abs(total_assets_end - total_liab_eq_end))),
        "ok": bs_ok_begin and bs_ok_end,
    })
    # 校验2：主表与补充资料经营净额一致
    supp_ok = abs(supp_op_net - op_net) < 0.01
    validations.append({
        "name": "经营活动现金流量净额（主表=补充资料）",
        "detail": "主表(%s) vs 补充资料(%s)" % (fmt(op_net), fmt(supp_op_net)),
        "ok": supp_ok,
    })
    # 校验3：现金及现金等价物净增加额勾稽
    # 高精度模式下主表净额=各项真实值求和，与货币资金变动天然不等（资金复杂企业），降级为提示
    cash_ok = abs(total_net - cash_increase) < 0.01
    cash_diff = abs(total_net - cash_increase)
    validations.append({
        "name": "现金及现金等价物净增加额（=货币资金及现金等价物变动）"
                + ("【提示】" if _other_op_in_override and not cash_ok else ""),
        "detail": "主表(%s) vs 货币资金变动(%s)%s"
                 % (fmt(total_net), fmt(cash_increase),
                    "（高精度模式：差异为资金复杂企业货币资金≠现金等价物所致，属正常）"
                    if _other_op_in_override and not cash_ok else ""),
        "ok": cash_ok or (_other_op_in_override and cash_diff > 0),
    })
    # 校验4：净利润与未分配利润变动勾稽（近似，提示用）
    undist = bs_get("未分配利润", "期末数") - bs_get("未分配利润", "年初数")
    validations.append({
        "name": "未分配利润变动 vs 净利润（提示）",
        "detail": "未分配利润增加(%s) vs 本年净利润(%s)；差异通常为利润分配，属正常"
                 % (fmt(undist), fmt(net_profit)),
        "ok": True,
    })

    result = {
        "main": {
            "销售商品提供劳务收到的现金": sale_cash,
            "收到的税费返还": tax_refund,
            "收到的其他与经营活动有关的现金": other_op_in,
            "经营现金流入小计": op_inflow,
            "购买商品接受劳务支付的现金": purchase_cash,
            "支付给职工以及为职工支付的现金": wage_paid,
            "支付的各项税费": tax_paid,
            "支付的其他与经营活动有关的现金": other_op_paid,
            "经营现金流出小计": op_outflow,
            "经营活动产生的现金流量净额": op_net,
            "收回投资所收到的现金": inv_retrieve,
            "取得投资收益所收到的现金": inv_income,
            "处置固定资产无形资产和其他长期资产收到的现金净额": inv_dispose,
            "收到的其他与投资活动有关的现金": inv_other_in,
            "投资现金流入小计": inv_inflow,
            "购建固定资产无形资产和其他长期资产支付的现金": fa_purchase,
            "投资所支付的现金": inv_pay,
            "支付的其他与投资活动有关的现金": inv_other_pay,
            "投资现金流出小计": inv_outflow,
            "投资活动产生的现金流量净额": inv_net3,
            "吸收投资所收到的现金": fin_absorb,
            "借款所收到的现金": fin_borrow,
            "收到的其他与筹资活动有关的现金": fin_other_in,
            "筹资现金流入小计": fin_inflow,
            "偿还债务所支付的现金": fin_repay,
            "分配股利利润或偿付利息所支付的现金": fin_dividend,
            "支付的其他与筹资活动有关的现金": fin_other_pay,
            "筹资现金流出小计": fin_outflow,
            "筹资活动产生的现金流量净额": fin_net3,
            "汇率变动对现金的影响": fx_change,
            "现金及现金等价物净增加额": total_net,
        },
        "supp": {
            "净利润": net_profit,
            "计提的资产减值准备": bad_debt,
            "固定资产折旧": dep,
            "无形资产摊销": intang_amort,
            "长期待摊费用摊销": lta_amort,
            "待摊费用减少": prepaid_dec,
            "预提费用增加": accrued_inc,
            "处置长期资产损失": dispose_loss,
            "固定资产报废损失": scrap_loss,
            "财务费用": fin_expense,
            "投资损失": inv_loss,
            "递延税款贷项": deferred_tax,
            "存货的减少": inv_dec,
            "经营性应收项目的减少": recv_dec,
            "经营性应付项目的增加": pay_inc,
            "其他": other_supp,
            "经营活动产生的现金流量净额": supp_op_net,
        },
        "cash": {
            "现金期末余额": cash_end,
            "现金期初余额": cash_begin,
            "现金等价物期末余额": eqv_end,
            "现金等价物期初余额": eqv_begin,
            "现金及现金等价物净增加额": cash_increase,
        },
        "adjustments": adjustments,
        "notes": notes,
        "validations": validations,
        "params_used": params,
    }
    return result


# ---------------------------------------------------------------------------
# 格式化与输出
# ---------------------------------------------------------------------------

def fmt(v, nd=2):
    """金额格式化：千分位，保留2位小数"""
    try:
        return f"{float(v):,.2f}"
    except (TypeError, ValueError):
        return "0.00"


def build_main_rows(result, prior=None):
    """构建主表输出行：[行次, 项目, 本期金额, 上期金额]"""
    m = result["main"]
    key_map = {
        "销售商品提供劳务收到的现金": "销售商品提供劳务收到的现金",
        "收到的税费返还": "收到的税费返还",
        "收到的其他与经营活动有关的现金": "收到的其他与经营活动有关的现金",
        "经营现金流入小计": "经营现金流入小计",
        "购买商品接受劳务支付的现金": "购买商品接受劳务支付的现金",
        "支付给职工以及为职工支付的现金": "支付给职工以及为职工支付的现金",
        "支付的各项税费": "支付的各项税费",
        "支付的其他与经营活动有关的现金": "支付的其他与经营活动有关的现金",
        "经营现金流出小计": "经营现金流出小计",
        "经营活动产生的现金流量净额": "经营活动产生的现金流量净额",
        "收回投资所收到的现金": "收回投资所收到的现金",
        "取得投资收益所收到的现金": "取得投资收益所收到的现金",
        "处置固定资产无形资产和其他长期资产收到的现金净额": "处置固定资产无形资产和其他长期资产收到的现金净额",
        "收到的其他与投资活动有关的现金": "收到的其他与投资活动有关的现金",
        "投资现金流入小计": "投资现金流入小计",
        "购建固定资产无形资产和其他长期资产支付的现金": "购建固定资产无形资产和其他长期资产支付的现金",
        "投资所支付的现金": "投资所支付的现金",
        "支付的其他与投资活动有关的现金": "支付的其他与投资活动有关的现金",
        "投资现金流出小计": "投资现金流出小计",
        "投资活动产生的现金流量净额": "投资活动产生的现金流量净额",
        "吸收投资所收到的现金": "吸收投资所收到的现金",
        "借款所收到的现金": "借款所收到的现金",
        "收到的其他与筹资活动有关的现金": "收到的其他与筹资活动有关的现金",
        "筹资现金流入小计": "筹资现金流入小计",
        "偿还债务所支付的现金": "偿还债务所支付的现金",
        "分配股利利润或偿付利息所支付的现金": "分配股利利润或偿付利息所支付的现金",
        "支付的其他与筹资活动有关的现金": "支付的其他与筹资活动有关的现金",
        "筹资现金流出小计": "筹资现金流出小计",
        "筹资活动产生的现金流量净额": "筹资活动产生的现金流量净额",
        "汇率变动对现金的影响": "汇率变动对现金的影响",
        "现金及现金等价物净增加额": "现金及现金等价物净增加额",
    }
    # 主表固定顺序（label, key, kind, 行次）
    order = [
        ("一、经营活动产生的现金流量：", None, "header", 1),
        ("  销售商品、提供劳务收到的现金", "销售商品提供劳务收到的现金", "item", 1),
        ("  收到的税费返还", "收到的税费返还", "item", 3),
        ("  收到的其他与经营活动有关的现金", "收到的其他与经营活动有关的现金", "item", 8),
        ("  现金流入小计", "经营现金流入小计", "sub", 9),
        ("  购买商品、接受劳务支付的现金", "购买商品接受劳务支付的现金", "item", 10),
        ("  支付给职工以及为职工支付的现金", "支付给职工以及为职工支付的现金", "item", 12),
        ("  支付的各项税费", "支付的各项税费", "item", 13),
        ("  支付的其他与经营活动有关的现金", "支付的其他与经营活动有关的现金", "item", 18),
        ("  现金流出小计", "经营现金流出小计", "sub", 20),
        ("  经营活动产生的现金流量净额", "经营活动产生的现金流量净额", "sub", 21),
        ("二、投资活动产生的现金流量：", None, "header", 22),
        ("  收回投资所收到的现金", "收回投资所收到的现金", "item", 22),
        ("  取得投资收益所收到的现金", "取得投资收益所收到的现金", "item", 23),
        ("  处置固定资产、无形资产和其他长期资产而收到的现金净额", "处置固定资产无形资产和其他长期资产收到的现金净额", "item", 25),
        ("  收到的其他与投资活动有关的现金", "收到的其他与投资活动有关的现金", "item", 28),
        ("  现金流入小计", "投资现金流入小计", "sub", 29),
        ("  购建固定资产、无形资产和其他长期资产所支付的现金", "购建固定资产无形资产和其他长期资产支付的现金", "item", 30),
        ("  投资所支付的现金", "投资所支付的现金", "item", 31),
        ("  支付的其他与投资活动有关的现金", "支付的其他与投资活动有关的现金", "item", 35),
        ("  现金流出小计", "投资现金流出小计", "sub", 36),
        ("  投资活动产生的现金流量净额", "投资活动产生的现金流量净额", "sub", 37),
        ("三、筹资活动产生的现金流量：", None, "header", 38),
        ("  吸收投资所收到的现金", "吸收投资所收到的现金", "item", 38),
        ("  借款所收到的现金", "借款所收到的现金", "item", 40),
        ("  收到的其他与筹资活动有关的现金", "收到的其他与筹资活动有关的现金", "item", 43),
        ("  现金流入小计", "筹资现金流入小计", "sub", 44),
        ("  偿还债务所支付的现金", "偿还债务所支付的现金", "item", 45),
        ("  分配股利、利润或偿付利息所支付的现金", "分配股利利润或偿付利息所支付的现金", "item", 46),
        ("  支付的其他与筹资活动有关的现金", "支付的其他与筹资活动有关的现金", "item", 52),
        ("  现金流出小计", "筹资现金流出小计", "sub", 53),
        ("  筹资活动产生的现金流量净额", "筹资活动产生的现金流量净额", "sub", 54),
        ("四、汇率变动对现金的影响", "汇率变动对现金的影响", "item", 55),
        ("五、现金及现金等价物净增加额", "现金及现金等价物净增加额", "sub", 56),
    ]
    rows = []
    for label, key, kind, lineno in order:
        this_v = m.get(key) if key else None
        prior_v = None
        if prior is not None and key:
            prior_v = prior["main"].get(key)
        rows.append({
            "行次": lineno,
            "项目": label,
            "本期金额": fmt(this_v) if this_v is not None else "",
            "上期金额": fmt(prior_v) if prior_v is not None else "",
            "kind": kind,
        })
    return rows


def build_supp_rows(result, prior=None):
    """构建补充资料输出行"""
    s = result["supp"]
    # (项目键, 显示名, 行次)
    order = [
        ("净利润", "  净利润", 57),
        ("计提的资产减值准备", "  加：计提的资产减值准备", 58),
        ("固定资产折旧", "        固定资产折旧", 59),
        ("无形资产摊销", "        无形资产摊销", 60),
        ("长期待摊费用摊销", "        长期待摊费用摊销", 61),
        ("待摊费用减少", "        待摊费用减少（减：增加）", 64),
        ("预提费用增加", "        预提费用增加（减：减少）", 65),
        ("处置长期资产损失", "        处置固定资产、无形资产和其他长期资产的损失（减：收益）", 66),
        ("固定资产报废损失", "        固定资产报废损失", 67),
        ("财务费用", "        财务费用", 68),
        ("投资损失", "        投资损失（减：收益）", 69),
        ("递延税款贷项", "        递延税款贷项（减：借项）", 70),
        ("存货的减少", "        存货的减少（减：增加）", 71),
        ("经营性应收项目的减少", "        经营性应收项目的减少（减：增加）", 72),
        ("经营性应付项目的增加", "        经营性应付项目的增加（减：减少）", 73),
        ("其他", "        其他", 74),
        ("经营活动产生的现金流量净额", "  经营活动产生的现金流量净额", 75),
    ]
    rows = []
    for name, disp, lineno in order:
        this_v = s.get(name)
        prior_v = None
        if prior is not None:
            prior_v = prior["supp"].get(name)
        rows.append({
            "行次": lineno,
            "项目": disp,
            "本期金额": fmt(this_v) if this_v is not None else "",
            "上期金额": fmt(prior_v) if prior_v is not None else "",
            "kind": "supp",
        })
    return rows


def to_markdown(result, title="现金流量表", prior=None):
    """生成 Markdown 报表"""
    lines = []
    lines.append(f"# {title}")
    lines.append("")
    lines.append("（依据资产负债表、利润及利润分配表及表外数据，按《现金流量表编制说明》口径自动生成）")
    lines.append("")
    lines.append("## 一、现金流量表（主表）")
    lines.append("")
    lines.append("| 行次 | 项目 | 本期金额 | 上期金额 |")
    lines.append("| ---: | --- | ---: | ---: |")
    for r in build_main_rows(result, prior):
        if r["kind"] == "header":
            lines.append(f"| {r['行次']} | **{r['项目']}** |  |  |")
        else:
            lines.append(f"| {r['行次']} | {r['项目']} | {r['本期金额']} | {r['上期金额']} |")
    lines.append("")
    lines.append("## 二、补充资料（间接法：净利润调节为经营活动现金流量）")
    lines.append("")
    lines.append("| 行次 | 项目 | 本期金额 | 上期金额 |")
    lines.append("| ---: | --- | ---: | ---: |")
    for r in build_supp_rows(result, prior):
        lines.append(f"| {r['行次']} | {r['项目']} | {r['本期金额']} | {r['上期金额']} |")
    lines.append("")
    lines.append("## 三、现金及现金等价物")
    lines.append("")
    lines.append("| 行次 | 项目 | 本期金额 | 上期金额 |")
    lines.append("| ---: | --- | ---: | ---: |")
    c = result["cash"]
    cash_key_map = {
        79: "现金期末余额", 80: "现金期初余额", 81: "现金等价物期末余额",
        82: "现金等价物期初余额", 83: "现金及现金等价物净增加额",
    }
    for ln, name in CASH_ITEMS:
        key = cash_key_map[ln]
        lines.append(f"| {ln} | {name} | {fmt(c.get(key, 0))} |  |")
    lines.append("")
    # 负数调整
    if result["adjustments"]:
        lines.append("## 四、负数调整说明（编制说明第四条）")
        lines.append("")
        for a in result["adjustments"]:
            lines.append(f"- {a}")
        lines.append("")
    # 校验
    lines.append("## 五、勾稽关系校验")
    lines.append("")
    for v in result["validations"]:
        lines.append(f"- {'✓' if v['ok'] else '✗'} **{v['name']}**：{v['detail']}")
    lines.append("")
    # 附注
    lines.append("## 六、编制附注（主要计算过程）")
    lines.append("")
    for n in result["notes"]:
        lines.append(f"- {n}")
    lines.append("")
    lines.append("> 说明：本表依据资产负债表与损益表编制，反映现金流量大致状况。"
                 "表外数据中的税率、坏账计提比例等须按公司实际情况核实；"
                 "涉及明细账簿的项目（如大额处置损益、非货币交易等）建议进一步核对。")
    return "\n".join(lines)


def to_excel(result, out_path, title="现金流量表"):
    """输出规范 Excel 报表（主表+补充资料+附注）"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    # ---- 主表 ----
    ws = wb.active
    ws.title = "现金流量表"
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_font = Font(name="宋体", size=14, bold=True)
    sub_font = Font(name="宋体", size=11, bold=True)
    normal_font = Font(name="宋体", size=11)
    head_fill = PatternFill("solid", fgColor="DDEBF7")
    header_fill = PatternFill("solid", fgColor="BDD7EE")
    net_fill = PatternFill("solid", fgColor="FCE4D6")

    ws.merge_cells("A1:D1")
    ws["A1"] = title
    ws["A1"].font = head_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A2:D2")
    ws["A2"] = "编制单位：＿＿＿＿＿＿　　年度：＿＿＿＿　　单位：元"
    ws["A2"].font = normal_font

    hdr = ["行次", "项目", "本期金额", "上期金额"]
    for ci, h in enumerate(hdr, 1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.font = sub_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    r = 4
    for row in build_main_rows(result):
        for ci, val in enumerate([row["行次"], row["项目"], row["本期金额"], row["上期金额"]], 1):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.font = normal_font
            cell.border = border
            if ci == 1:
                cell.alignment = Alignment(horizontal="center")
            if ci == 2:
                cell.alignment = Alignment(horizontal="left", indent=1)
                if row["kind"] == "header":
                    cell.font = sub_font
                    cell.fill = head_fill
            if ci in (3, 4):
                cell.alignment = Alignment(horizontal="right")
                if row["kind"] == "sub":
                    cell.font = Font(name="宋体", size=11, bold=True)
        if row["kind"] == "sub":
            for ci in range(1, 5):
                ws.cell(row=r, column=ci).fill = net_fill
        r += 1

    widths = [8, 52, 16, 16]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ---- 补充资料 ----
    ws2 = wb.create_sheet("补充资料")
    ws2.merge_cells("A1:D1")
    ws2["A1"] = "补充资料（间接法）"
    ws2["A1"].font = head_font
    ws2["A1"].alignment = Alignment(horizontal="center")
    for ci, h in enumerate(hdr, 1):
        cell = ws2.cell(row=3, column=ci, value=h)
        cell.font = sub_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
    r = 4
    for row in build_supp_rows(result):
        for ci, val in enumerate([row["行次"], row["项目"], row["本期金额"], row["上期金额"]], 1):
            cell = ws2.cell(row=r, column=ci, value=val)
            cell.font = normal_font
            cell.border = border
            cell.alignment = Alignment(horizontal="right" if ci in (3, 4) else ("center" if ci == 1 else "left"))
        r += 1
    for ci, w in enumerate(widths, 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    # ---- 编制附注 ----
    ws3 = wb.create_sheet("编制附注与校验")
    ws3.merge_cells("A1:F1")
    ws3["A1"] = "编制附注与勾稽校验"
    ws3["A1"].font = head_font
    r = 3
    ws3.cell(row=r, column=1, value="【勾稽关系校验】").font = sub_font
    r += 1
    for v in result["validations"]:
        ws3.cell(row=r, column=1, value=f"{'✓' if v['ok'] else '✗'} {v['name']}：{v['detail']}")
        r += 1
    if result["adjustments"]:
        r += 1
        ws3.cell(row=r, column=1, value="【负数调整说明】").font = sub_font
        r += 1
        for a in result["adjustments"]:
            ws3.cell(row=r, column=1, value=f"- {a}")
            r += 1
    r += 1
    ws3.cell(row=r, column=1, value="【主要计算过程】").font = sub_font
    r += 1
    for n in result["notes"]:
        ws3.cell(row=r, column=1, value=f"- {n}")
        r += 1
    ws3.column_dimensions["A"].width = 140
    wb.save(out_path)


# ---------------------------------------------------------------------------
# 主流程
# ---------------------------------------------------------------------------

def run(bs_data=None, pl_data=None, extra_data=None,
        bs_prior_data=None, pl_prior_data=None, params=None):
    """完整流程：解析 → 计算 → 返回结果。
    若只传入一个 Excel 工作簿（含资产负债表/利润表/表外数据多个工作表），
    自动从同一工作簿识别各表；单独提供的 pl/extra 优先。
    """
    params = params or DEFAULT_PARAMS

    bs_src = load_table(bs_data)
    pl_src = load_table(pl_data) if pl_data is not None else None
    extra_src = load_table(extra_data) if extra_data is not None else None

    # 未单独提供利润表/表外数据时，若主输入是工作簿（含多个工作表），复用同一数据源
    if pl_src is None and bs_src is not None and len(bs_src) > 1:
        pl_src = bs_src
    if extra_src is None and bs_src is not None and len(bs_src) > 1:
        extra_src = bs_src

    bs = parse_balance_sheet(bs_src)
    pl = parse_income_statement(pl_src)
    extra = parse_extra_data(extra_src, bs, pl, params)

    result = compute_cash_flow(bs, pl, extra, params)

    prior = None
    if bs_prior_data is not None or pl_prior_data is not None:
        bs_p_src = load_table(bs_prior_data)
        pl_p_src = load_table(pl_prior_data) if pl_prior_data is not None else None
        if pl_p_src is None and bs_p_src is not None and len(bs_p_src) > 1:
            pl_p_src = bs_p_src
        extra_p_src = load_table(extra_data) if extra_data is not None else extra_src
        bs_p = parse_balance_sheet(bs_p_src)
        pl_p = parse_income_statement(pl_p_src)
        extra_p = parse_extra_data(extra_p_src, bs_p, pl_p, params)
        prior = compute_cash_flow(bs_p, pl_p, extra_p, params)
    return result, prior


def main():
    ap = argparse.ArgumentParser(description="现金流量表自动编制引擎")
    ap.add_argument("--bs", required=True, help="资产负债表文件（xlsx/xls/csv 或 Markdown 表格文本文件）；若为含多表的工作簿可单独使用")
    ap.add_argument("--pl", help="利润及利润分配表文件（工作簿模式下可省略）")
    ap.add_argument("--extra", help="表外数据文件（工作簿模式下可省略）")
    ap.add_argument("--bs-prior", help="上年资产负债表（用于上期金额）")
    ap.add_argument("--pl-prior", help="上年利润表")
    ap.add_argument("--out", default=".", help="输出目录")
    ap.add_argument("--fmt", choices=["markdown", "excel", "json", "all"], default="all")
    ap.add_argument("--json-params", help='覆盖参数 JSON，如 {"sale_vat_rate":0.13}')
    args = ap.parse_args()

    params = dict(DEFAULT_PARAMS)
    if args.json_params:
        params.update(json.loads(args.json_params))

    result, prior = run(args.bs, args.pl, args.extra,
                        args.bs_prior, args.pl_prior, params)

    os.makedirs(args.out, exist_ok=True)
    base = os.path.join(args.out, "现金流量表")
    if args.fmt in ("markdown", "all"):
        md = to_markdown(result, prior=prior)
        print(md)
        with open(base + ".md", "w", encoding="utf-8") as f:
            f.write(md)
    if args.fmt in ("excel", "all"):
        to_excel(result, base + ".xlsx")
    if args.fmt in ("json", "all"):
        with open(base + ".json", "w", encoding="utf-8") as f:
            json.dump({"result": result, "prior": prior}, f, ensure_ascii=False,
                      indent=2, default=str)
    print(f"\n[输出] {base}.md / .xlsx / .json（已生成）", file=sys.stderr)


if __name__ == "__main__":
    main()
