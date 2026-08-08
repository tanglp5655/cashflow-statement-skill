# -*- coding: utf-8 -*-
"""重新生成《表外数据收集表.xlsx / .csv》到 examples/ 目录。

设计要点（贴近原模板 编制现金流量表2024.xls）：
1. **粉红阴影 = 公式自动算**（合计 / 实际应缴纳各项税金合计 / 分配股利、利润或偿付利息所支付的现金）
   用户看到粉红色单元格就知道是自动算的，不要手填。
2. **税率列**给销项/进项/所得税/坏账计提比例填默认值（与原模板一致），用户按公司实际改。
3. **顶部警告**醒目提示"白色空格录入数据，粉红色公式格不要填"。
4. 引擎通过 _cell_num() 拿不到数字时自动跳过，**公式行不影响引擎运行**——引擎自身在
   compute_cash_flow 里也会算这些合计值（更准确，包含资产负债表跨表勾稽）。

用法：python scripts/gen_extra_form.py
"""
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
OUT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "examples")

# ---------------------------------------------------------------------------
# 第一区：编制现金流量表所需表外数据
# 格式：(kind, 项目, 说明, 必填/选填, 税率默认, 税率说明)
# kind: "DATA" = 录入项；"FORMULA" = 公式自动算（粉红色阴影）
# 公式项写在它要汇总的 DATA 项之后，引擎用 row 编号写公式。
# ---------------------------------------------------------------------------
SECTION1 = [
    # 工资块
    ("DATA", "应收票据贴现利息支出", "票据贴现产生的利息（如无贴现可留空）", "选填", None, None),
    ("DATA", "★ 支付给职工的工资", "全年实发工资总额。系统不自动推算，必须手填（直接影响经营净额）", "必填", None, None),
    ("DATA", "支付给职工的四金", "社保+公积金等。留空时系统按工资×26.6%自动推算", "选填", None, None),
    ("DATA", "支付给职工的其他福利费", "其他职工福利。留空时系统按工资×1.06%自动推算", "选填", None, None),
    # ===== 公式项1：合计 = 工资+四金+福利费 =====
    ("FORMULA", "⚡ 合计（工资+四金+福利费）", None, "自动算", None, None, "=C{r1}+C{r2}+C{r3}", "(公式：工资+四金+福利费)"),

    # 税费块
    ("DATA", "销项税额", "全年销项税额。留空时系统按收入×6%推算（税率默认6%，须按实际改）", "选填", 0.06, "须根据实际税率修改"),
    ("DATA", "进项税额", "全年进项税额。留空时系统按成本×6%推算", "选填", 0.00, "须根据实际税率修改"),
    # ===== 公式项1b：应交增值税 = 销项 - 进项（自动算）
    ("FORMULA", "⚡ 应交增值税", None, "自动算", None, None,
     "=C{r2}-C{r1}", "(公式：销项税额-进项税额)"),
    ("DATA", "其他各项税", "除增值税外各税种（城建税/教育附加等）。留空时系统取利润表营业税金及附加", "选填", None, None),
    ("DATA", "所得税", "全年所得税。留空时系统取利润表所得税费用", "选填", 0.25, "须根据实际税率修改"),
    ("DATA", "管理费用中列支的税金", "计入管理费用的税金（房产税、印花税、车船税等）", "选填", None, None),
    ("DATA", "在其他业务支出中列支的税金", "其他业务负担的税金", "选填", None, None),
    # ===== 公式项2：实际应缴纳各项税金合计 =====
    # = 应交增值税 + 其他各项税 + 所得税 + 管理税金 + 其他业务税金
    # 注：原模板还减 应交税费/其他应交款 期末-期初 变动（涉及跨表引用）；引擎在 compute_cash_flow
    # 里会按更完整口径（含跨表勾稽）算 actual_tax_paid，所以模板里简化展示。
    ("FORMULA", "⚡ 实际应缴纳各项税金合计", None, "自动算", None, None,
     "=C{r1}+C{r2}+C{r3}+C{r4}+C{r5}", "(公式：其他业务税金+管理税金+所得税+其他各项税+应交增值税)"),

    # 股利/利润/利息块
    ("DATA", "分配股利所支付的现金", "全年向股东分配股利支付的现金", "选填", None, None),
    ("DATA", "分配利润所支付的现金", "全年分配利润支付的现金", "选填", None, None),
    ("DATA", "利息支出", "利息支出（不抵减利息收入；利润表财务费用中的利息部分）", "选填", None, None),
    # ===== 公式项3：分配股利、利润或偿付利息所支付的现金 =====
    ("FORMULA", "⚡ 分配股利、利润或偿付利息所支付的现金", None, "自动算", None, None,
     "=C{r1}+C{r2}+C{r3}", "(公式：利息支出+分配利润+分配股利)"),

    # 坏账/折旧摊销/处置块
    ("DATA", "计提坏账准备的比率", "坏账计提比例（默认0.5%，按公司实际改）", "选填", 0.005, "可根据实际计提比率修改"),
    ("DATA", "计提的坏账准备", "当年计提坏账（不填为0，与模板口径一致）", "选填", None, None),
    ("DATA", "无形资产摊销", "当年无形资产摊销额", "选填", None, None),
    ("DATA", "长期待摊费用摊销", "当年长期待摊费用摊销额", "选填", None, None),
    ("DATA", "固定资产报废损失", "当年报废损失（不计入处置现金净额）", "选填", None, None),
    ("DATA", "收到投资分红或利润", "当年收到的对外投资分红/利润", "选填", None, None),
    ("DATA", "处置固定资产收回的现金净额", "处置固定资产收到的现金净额", "选填", None, None),
    ("DATA", "处置无形资产收回的现金净额", "处置无形资产收到的现金净额", "选填", None, None),
    ("DATA", "处置其他长期资产收回的现金净额", "处置其他长期资产收到的现金净额", "选填", None, None),
    ("DATA", "收到的其他与投资活动有关的现金", "投资活动其他现金流入", "选填", None, None),
    ("DATA", "支付的其他与投资活动有关的现金", "投资活动其他现金流出", "选填", None, None),
    ("DATA", "收到的其他与筹资活动有关的现金", "筹资活动其他现金流入", "选填", None, None),
    ("DATA", "偿还债务所支付的现金", "偿还债务本金支付的现金（不填则按借款变动自动倒挤）", "选填", None, None),
    ("DATA", "支付的其他与筹资活动有关的现金", "筹资活动其他现金流出", "选填", None, None),
    ("DATA", "汇率变动对现金的影响", "外币折算对现金的影响（无外币业务填0）", "选填", None, None),
    ("DATA", "现金等价物期末余额", "如有现金等价物（一般企业可留空）", "选填", None, None),
    ("DATA", "现金等价物期初余额", "如有现金等价物（一般企业可留空）", "选填", None, None),
    ("DATA", "支付的其他与经营活动有关的现金", "经营活动其他现金流出（不填则由平衡项自动倒挤）", "选填", None, None),
]

# ===== 第二区：高精度模式（可选）：从现金流量表主表直接抄真实值 =====
# 填了这些项后，引擎用真实值替代公式推算，误差可压到 ±1% 以内
HIGH_PRECISION_ITEMS = [
    ("销售商品提供劳务收到的现金", '现金流量表主表"销售商品、提供劳务收到的现金"（高精度模式，选填）', "选填"),
    ("收到的税费返还", '现金流量表主表"收到的税费返还"（高精度模式，选填）', "选填"),
    ("购买商品接受劳务支付的现金", '现金流量表主表"购买商品、接受劳务支付的现金"（高精度模式，选填）', "选填"),
    ("支付给职工以及为职工支付的现金", '现金流量表主表"支付给职工以及为职工支付的现金"（高精度模式，选填）', "选填"),
    ("支付的各项税费", '现金流量表主表"支付的各项税费"（高精度模式，选填）', "选填"),
    ("收到的其他与经营活动有关的现金", '现金流量表主表"收到的其他与经营活动有关的现金"（高精度模式，选填）★填此项后经营净额不再倒挤，按各项真实值求和', "选填"),
    ("支付的其他与经营活动有关的现金", '现金流量表主表"支付的其他与经营活动有关的现金"（高精度模式，选填）', "选填"),
    ("购建固定资产无形资产和其他长期资产支付的现金", '现金流量表主表"购建固定资产、无形资产和其他长期资产所支付的现金"（高精度模式，选填）', "选填"),
    ("投资所支付的现金", '现金流量表主表"投资所支付的现金"（高精度模式，选填）', "选填"),
    ("吸收投资所收到的现金", '现金流量表主表"吸收投资收到的现金"（高精度模式，选填）', "选填"),
    ("借款所收到的现金", '现金流量表主表"借款收到的现金"（高精度模式，选填）', "选填"),
]

NOTES = [
    "★ = 必填项（系统不自动推算，必须手填）。⚡ = 自动算公式（粉红阴影，不要填）。其他 = 选填（留空时系统按说明自动推算）。",
    "★ 全年实发工资是唯一必填项；其余留空时系统自动按默认比例推算（详见每个项目的「说明」列）。",
    "⚙ 参数覆盖子表：四金比例/福利费比例/销项税率/进项税率/坏账计提比例——按公司实际修改，留空用模板默认。",
    "★ 税率列默认值：销项 6% / 进项 0% / 所得税 25% / 坏账 0.5%，请按公司实际修改。",
    "★ 数据来源列（浅蓝）：建议每个数字注明出处（如「管理费用明细-工资」），便于审计追溯。",
    "★ 填完后保存，运行时作为 --extra 参数传入：python cash_flow_generator.py --bs 资产负债表 --pl 利润表 --extra 表外数据收集表.xlsx",
    "★ 补齐本表后，「经营活动产生的现金流量净额」精度可由 ±20% 提升至 ±5% 以内。",
    "★ 数据来源建议：工资表/社保申报表、增值税及所得税申报表、固定资产折旧明细表、银行对账单。",
]


def _resolve_formulas(rows):
    """把 FORMULA 行的 {r1}/{r2}/... 占位符解析成 __ROW_k__ 中间标记。

    约定（直觉、易用）：{r1} = 紧邻 FORMULA 的上一行（最近的），
                          {r2} = 倒数第二行，...以此类推。
    注意：{rk} 引用的是"任意类型的行"（DATA + FORMULA），不是只引用 DATA。
    这样链式公式可以引用前置的 FORMULA 行（如"实际应缴纳合计"引用"应交增值税"）。
    gen_xlsx 时再把 __ROW_k__ 替换成 r-k（r 是 FORMULA 所在行号）。
    """
    resolved = []
    for i, row in enumerate(rows):
        if row[0] != "FORMULA":
            resolved.append(row)
            continue
        _, name, _, req, _, _, formula_tpl, note = row
        # 收集前面所有行（DATA + FORMULA）的个数
        n_above = i
        formula = formula_tpl
        for k in range(1, n_above + 1):
            placeholder = "{r" + str(k) + "}"
            formula = formula.replace(placeholder, f"__ROW_{k}__")
        resolved.append((row[0], name, None, req, None, None, formula, note, n_above))
    return resolved


def gen_xlsx(path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "表外数据收集表"

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_font = Font(name="宋体", size=14, bold=True)
    head_font = Font(name="宋体", size=11, bold=True)
    normal_font = Font(name="宋体", size=11)
    must_fill = PatternFill("solid", fgColor="FFF2CC")     # 必填：浅黄
    head_fill = PatternFill("solid", fgColor="BDD7EE")     # 表头：浅蓝
    note_fill = PatternFill("solid", fgColor="F2F2F2")     # 说明列：浅灰
    warn_fill = PatternFill("solid", fgColor="FFE699")     # 顶部警告：浅橙
    warn_font = Font(name="宋体", size=11, bold=True, color="C00000")
    formula_fill = PatternFill("solid", fgColor="FFC7CE") # 公式自动算：粉红
    formula_font = Font(name="宋体", size=11, bold=True, color="9C0006")
    rate_fill = PatternFill("solid", fgColor="E2EFDA")     # 税率：浅绿
    source_fill = PatternFill("solid", fgColor="DEEBF7")   # 数据来源（用户填）：浅蓝
    section_fill = PatternFill("solid", fgColor="305496") # 大区段标题：深蓝
    section_font = Font(name="宋体", size=12, bold=True, color="FFFFFF")
    hp_head_fill = PatternFill("solid", fgColor="C6EFCE") # 高精度区表头：浅绿

    # ---------- 顶部 ----------
    ws.merge_cells("A1:G1")
    ws["A1"] = "表 外 数 据 收 集 表"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ---------- 顶部警告（醒目）----------
    ws.merge_cells("A2:G2")
    ws["A2"] = ("⚠ 重要提醒：粉红阴影部分已设置为公式，数据自动计算产生，"
                "务请不要填入任何数据以免出错！只需在「金额」栏白色空格内录入数据。"
                "建议在「数据来源」栏填写每个数字的出处（如'管理费用明细表-工资'），方便审计追溯。")
    ws["A2"].font = warn_font
    ws["A2"].fill = warn_fill
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 48

    # ---------- 区段 1 标题 ----------
    ws.merge_cells("A3:G3")
    ws["A3"] = "一、编制现金流量表所需数据录入"
    ws["A3"].font = section_font
    ws["A3"].fill = section_fill
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 22

    # ---------- 区段 1 表头 ----------
    hdr1 = ["序号", "项  目", "金  额", "税  率", "税率说明", "数据来源", "说  明"]
    for ci, h in enumerate(hdr1, 1):
        c = ws.cell(row=4, column=ci, value=h)
        c.font = head_font
        c.fill = head_fill
        c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center")

    # ---------- 区段 1 数据 + 公式 ----------
    section1_resolved = _resolve_formulas(SECTION1)
    r = 5
    seq = 0
    for row_def in section1_resolved:
        kind = row_def[0]
        if kind == "DATA":
            _, name, note, req, rate, rate_note = row_def
            seq += 1
            ws.cell(row=r, column=1, value=seq).border = border
            ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")
            cname = ws.cell(row=r, column=2, value=name)
            cname.border = border
            cname.font = normal_font
            cval = ws.cell(row=r, column=3)
            cval.border = border
            cval.alignment = Alignment(horizontal="right")
            cval.font = normal_font
            crate = ws.cell(row=r, column=4)
            if rate is not None:
                crate.value = rate
                crate.number_format = "0.00%"
            crate.border = border
            crate.alignment = Alignment(horizontal="center")
            crate.fill = rate_fill
            crate.font = normal_font
            crate_note = ws.cell(row=r, column=5, value=rate_note or "")
            crate_note.border = border
            crate_note.font = Font(name="宋体", size=9, color="808080")
            crate_note.fill = note_fill
            # F 数据来源（用户填——空白单元格，浅蓝底鼓励填写）
            csource = ws.cell(row=r, column=6)
            csource.border = border
            csource.fill = source_fill
            csource.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            csource.font = Font(name="宋体", size=9, italic=True, color="666666")
            # G 说明（引擎提示，灰色）
            cnote = ws.cell(row=r, column=7, value=note)
            cnote.border = border
            cnote.font = Font(name="宋体", size=9, color="808080")
            cnote.fill = note_fill
            cnote.alignment = Alignment(wrap_text=True, vertical="center")
            if req == "必填":
                cname.fill = must_fill
                cval.fill = must_fill
            r += 1
        else:  # FORMULA
            _, name, _note_unused, req, _, _, formula_tpl, note, n_above = row_def
            formula = formula_tpl
            for k in range(1, n_above + 1):
                abs_row = r - k
                formula = formula.replace(f"__ROW_{k}__", str(abs_row))
            ws.cell(row=r, column=1).border = border
            ws.cell(row=r, column=1).fill = formula_fill
            cname = ws.cell(row=r, column=2, value=name)
            cname.border = border
            cname.font = formula_font
            cname.fill = formula_fill
            cval = ws.cell(row=r, column=3, value=formula)
            cval.border = border
            cval.alignment = Alignment(horizontal="right")
            cval.font = formula_font
            cval.fill = formula_fill
            cval.number_format = "#,##0.00;-#,##0.00;-"
            crate = ws.cell(row=r, column=4)
            crate.border = border
            crate.fill = formula_fill
            crate_note = ws.cell(row=r, column=5)
            crate_note.border = border
            crate_note.fill = formula_fill
            csource = ws.cell(row=r, column=6)
            csource.border = border
            csource.fill = formula_fill
            cnote = ws.cell(row=r, column=7, value=note)
            cnote.border = border
            cnote.font = Font(name="宋体", size=9, color="9C0006")
            cnote.fill = formula_fill
            cnote.alignment = Alignment(wrap_text=True, vertical="center")
            r += 1

    # ---------- 区段 2 标题（高精度模式）----------
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    hcell = ws.cell(row=r, column=1, value="二、【高精度模式·可选】从现金流量表主表直接抄真实值")
    hcell.font = section_font
    hcell.fill = PatternFill("solid", fgColor="548235")
    hcell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 22
    r += 1

    # ---------- 区段 2 表头 ----------
    hdr2 = ["序号", "项  目", "金  额", "", "", "数据来源", "说  明"]
    for ci, h in enumerate(hdr2, 1):
        c = ws.cell(row=r, column=ci, value=h)
        c.font = head_font
        c.fill = hp_head_fill
        c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center")
    # 合并 D:E（高精度模式没税率列）
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
    r += 1

    # ---------- 区段 2 数据 ----------
    for i, (name, note, req) in enumerate(HIGH_PRECISION_ITEMS, 1):
        ws.cell(row=r, column=1, value=i).border = border
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")
        cname = ws.cell(row=r, column=2, value=name)
        cname.border = border
        cname.font = normal_font
        cval = ws.cell(row=r, column=3)
        cval.border = border
        cval.alignment = Alignment(horizontal="right")
        cval.font = normal_font
        # D:E 合并留白（无税率）
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
        merged = ws.cell(row=r, column=4)
        merged.border = border
        merged.fill = note_fill
        # F 数据来源（用户填）
        csource = ws.cell(row=r, column=6)
        csource.border = border
        csource.fill = source_fill
        csource.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        # G 说明
        cnote = ws.cell(row=r, column=7, value=note)
        cnote.border = border
        cnote.font = Font(name="宋体", size=9, color="808080")
        cnote.fill = note_fill
        cnote.alignment = Alignment(wrap_text=True, vertical="center")
        r += 1

    # ---------- 区段 1.5：⚙ 系统默认参数覆盖 ----------
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    title = ws.cell(row=r, column=1, value="⚙ 系统默认参数覆盖（按公司实际修改，留空用默认）")
    title.font = Font(name="宋体", size=11, bold=True, color="FFFFFF")
    title.fill = PatternFill("solid", fgColor="4472C4")
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 22
    r += 1
    # 表头
    hdr_param = ["", "参  数  名", "实际值(0~1)", "默认值", "参数说明", "数据来源(选填)", "说  明"]
    for ci, h in enumerate(hdr_param, 1):
        c = ws.cell(row=r, column=ci, value=h)
        c.font = head_font
        c.fill = PatternFill("solid", fgColor="D9E1F2")
        c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center")
    r += 1
    # 参数定义：(参数名, 默认值, 0~1 范围说明, 详细说明)
    PARAMS = [
        ("⚙ 四金比例", 0.266, "社保+公积金等占工资比。各地差异大：上海约34%/北京约40%/深圳约26%/最低基数10-15%",
         "支付给职工的四金(留空时) = 工资 × 此比例"),
        ("⚙ 福利费比例", 0.0106, "其他职工福利费占工资比（通常1-2%）",
         "支付给职工的其他福利费(留空时) = 工资 × 此比例"),
        ("⚙ 销项税率", 0.06, "增值税销项税率（一般纳税人6%/9%/13%按行业；小规模3%）",
         "销项税额(留空时) = 营业收入 × 此税率"),
        ("⚙ 进项税率", 0.06, "增值税进项税率（同销项税率原则）",
         "进项税额(留空时) = 营业成本 × 此税率"),
        ("⚙ 坏账计提比例", 0.005, "坏账计提占应收账款比例（一般企业0.5-1%）",
         "此比例供填表参考；坏账实际计提额在主表「计提的坏账准备」行录入"),
    ]
    for i, (pname, default, range_note, detail) in enumerate(PARAMS, 1):
        ws.cell(row=r, column=1, value=i).border = border
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")
        cname = ws.cell(row=r, column=2, value=pname)
        cname.border = border
        cname.font = Font(name="宋体", size=11, bold=True, color="0070C0")
        cval = ws.cell(row=r, column=3)
        # 默认填上模板默认值，操作员可改
        cval.value = default
        cval.number_format = "0.00%"
        cval.border = border
        cval.alignment = Alignment(horizontal="right")
        cval.fill = PatternFill("solid", fgColor="FFF2CC")  # 浅黄突出可改
        cval.font = normal_font
        cdef = ws.cell(row=r, column=4, value=default)
        cdef.number_format = "0.00%"
        cdef.border = border
        cdef.alignment = Alignment(horizontal="center")
        cdef.font = Font(name="宋体", size=10, color="808080")
        cdef.fill = note_fill
        crate_note = ws.cell(row=r, column=5, value=range_note)
        crate_note.border = border
        crate_note.font = Font(name="宋体", size=9, color="808080")
        crate_note.fill = note_fill
        crate_note.alignment = Alignment(wrap_text=True, vertical="center")
        csrc = ws.cell(row=r, column=6)
        csrc.border = border
        csrc.fill = source_fill
        cnote = ws.cell(row=r, column=7, value=detail)
        cnote.border = border
        cnote.font = Font(name="宋体", size=9, color="808080")
        cnote.fill = note_fill
        cnote.alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[r].height = 32
        r += 1

    # ---------- 尾部说明 ----------
    r += 1
    for note in NOTES:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        cell = ws.cell(row=r, column=1, value="※ " + note)
        cell.font = Font(name="宋体", size=9, color="C00000")
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[r].height = 18
        r += 1

    # ---------- 列宽 ----------
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 32   # 数据来源（用户填）
    ws.column_dimensions["G"].width = 40   # 说明（引擎提示）

    wb.save(path)
    print("已生成:", path)


def gen_csv(path):
    """CSV 版本：纯文本，不带公式和颜色样式，只保留数据项。

    公式行在 CSV 里无法表达，跳过。"""
    import csv
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        # 警告 + 表头
        w.writerow(["⚠ 重要提醒：粉红阴影部分已设置为公式，数据自动计算产生，务请不要填入任何数据以免出错！"
                    "只需在「金额」栏白色空格内录入数据。建议在「数据来源」栏填每个数字的出处。"])
        w.writerow([])
        w.writerow(["一、编制现金流量表所需数据录入（公式项：合计 / 应交增值税 / 实际应缴纳各项税金合计 / "
                    "分配股利、利润或偿付利息所支付的现金 = 自动算，请勿手填）"])
        w.writerow(["序号", "项目", "金额", "税率(可选)", "税率说明", "数据来源(选填)", "说明"])
        seq = 0
        for row_def in SECTION1:
            if row_def[0] != "DATA":
                # FORMULA 行：在 CSV 里写一行说明性记录（空值 + 说明）
                _, name, _, req, _, _, _, note = row_def
                seq += 1  # 也给个序号方便对照
                w.writerow([seq, f"【公式自动算·勿填】{name}", "", "", "", "", note])
                continue
            _, name, note, req, rate, rate_note = row_def
            seq += 1
            rate_str = ""
            if rate is not None:
                rate_str = f"{rate*100:.2f}%" + (f" ({rate_note})" if rate_note else "")
            w.writerow([seq, name, "", rate_str, rate_note or "", "", note])
        w.writerow([])
        w.writerow(["二、【高精度模式·可选】从现金流量表主表直接抄真实值"])
        w.writerow(["序号", "项目", "金额", "数据来源(选填)", "说明"])
        for i, (name, note, _req) in enumerate(HIGH_PRECISION_ITEMS, 1):
            w.writerow([i, name, "", "", note])
        w.writerow([])
        for note in NOTES:
            w.writerow(["※ " + note])
    print("已生成:", path)


if __name__ == "__main__":
    os.makedirs(OUT_DIR, exist_ok=True)
    gen_xlsx(os.path.join(OUT_DIR, "表外数据收集表.xlsx"))
    gen_csv(os.path.join(OUT_DIR, "表外数据收集表.csv"))